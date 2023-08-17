from flask import Flask, flash, session, render_template,\
    request, redirect, url_for
from flask_session import Session
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from datetime import datetime, date
import pymysql.cursors

app = Flask(__name__)
app.secret_key = "Smansa"
#Folder untuk meletakkan file upload
upl = r'C:\Users\fenar\Smansa\static\uploads'
app.config['UPLOAD_FOLDER'] = upl
app.config['MAX_CONTENT_PATH'] = 10000000
#Session disimpan pada sistem dan tidak permanen
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

conn = cursor = None

# Deklarasi variabel global status berhasil/error untuk diakses beberapa fungsi
Stat = True
lanjut = True
# Variabel dummy untuk menguji jika gender = P maka perempuan, sebaliknya laki-laki
gend = "P"

#Akses database
def openDb():
    global conn, cursor
    conn = pymysql.connect(
        host= 'localhost',
        user= 'root',
        password= "",
        db= 'smansa'
    )
    cursor = conn.cursor()

#Keluar database
def closeDb():
    global conn, cursor
    cursor.close()
    conn.close()

#Verifikasi login berdasarkan hak akses
def verifLogin(a):
    if not 'login' in session: #Jika belum login
        logout()
        flash('Silahkan login terlebih dahulu')
        return False
    elif not session["akses"]==a: #Jika sudah login tapi beda akses
        logout()
        flash('Akun ini tidak memiliki akses')
        flash('Silahkan login kembali')
        return False
    return True

#Penamaan hari berdasarkan angka
def cekHari(dow):
    nmHr = ("","Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu")
    return nmHr[dow]

def cekHariR(hari):
    nmHr = ("","Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu")
    return nmHr.index(hari)

#Penamaan bulan berdasarkan angka
def cekBulan(mm):
    bln = ("","Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember")
    return bln[mm];

#Fungsi untuk membersihkan session ketika logout
@app.route('/logout')
def logout():
    session.pop('login', None)
    session.pop('nama', None)
    session.pop('id', None)
    session.pop('akses', None)
    return redirect('/')

#Halaman login
@app.route('/', methods = ['GET', 'POST'])
def index():
    #Jika user menekan tombol login
    if request.method == 'POST' and 'username' in request.form and 'password' in request.form:
        openDb()
        #Verifikasi login
        #Username bisa id atau email, bebas
        username = request.form['username']
        password = request.form['password']
        #Cek jika akun benar ada atau tidak
        sql = "SELECT * FROM user where (username = %s OR email = %s) AND password= %s"
        cursor.execute(sql, (username, username, password))
        login = cursor.fetchone()
        #Jika akun ditemukan dan cocok
        if login:
            #Simpan semua data user yang login ke session
            session['login'] = True
            session['nama'] = login[1]
            session['id'] = login[0]
            session['akses'] = login[4]
            #Login berhasil, lanjut ke pengecekan hak akses
            if login[4] == 1:
                return redirect(url_for('admin')) #Login admin
            elif login[4] == 2:
                return redirect(url_for('guru')) #Login guru
            elif login[4] == 3:
                return redirect(url_for('siswa')) #Login siswa
            elif login[4] == 4:
                return redirect(url_for('perpus')) #Login perpustakaan
            else:
                flash('Akun tidak valid') #Hak akses tidak ditemukan atau tidak valid, antisipasi error
                return redirect(request.url)
        #Login gagal karena username atau password kosong
        elif username == "" or password == "":
            flash('Username atau password tidak boleh kosong!')
        #Login gagal karena username atau password salah
        else:
            flash('Username atau password salah')
        closeDb()
    return render_template('index.html')


''' Akses Admin '''
@app.route('/dashboard_admin', methods = ['GET', 'POST'])
def admin():
    if verifLogin(1):
        openDb()
        # Fetch hari & tanggal saat ini
        tanggal = date.today()
        d = tanggal.strftime("%d")
        m = cekBulan(tanggal.month)
        y = tanggal.strftime("%Y")
        # dt = tanggal
        dt = d + " " + m + " " + y
        dow = tanggal.isoweekday()
        # numHari = hari
        numHari = cekHari(dow)

        # Fetch jumlah guru
        sql = "SELECT COUNT(*) FROM guru"
        cursor.execute(sql)
        res = cursor.fetchone()
        numGuru = res[0]

        # Fetch jumlah siswa
        sql = "SELECT COUNT(*) FROM siswa"
        cursor.execute(sql)
        res = cursor.fetchone()
        numSiswa = res[0]

        # Fetch jumlah kelas
        sql = "SELECT COUNT(*) FROM kelas"
        cursor.execute(sql)
        res = cursor.fetchone()
        numKelas = res[0]

        # Fetch jumlah staf
        sql = "SELECT COUNT(*) FROM user WHERE username NOT IN('admin') AND access = 1 OR access = 4"
        cursor.execute(sql)
        res = cursor.fetchone()
        numStaf = res[0]

        closeDb()
        return render_template('admin/dashboard_admin.html', guru=numGuru, siswa=numSiswa, hari=numHari, tanggal=dt, kelas=numKelas, staf=numStaf)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_guru', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_admin/data_guru/<int:page>', methods = ['GET', 'POST'])
def dataguru(page):
    if verifLogin(1):
        openDb()
        global Stat
        lanjut = True
        pencarian = ""
        prev = page-1
        next = page+1
        counter = prev*10

        #Ambil daftar guru dari database
        sql = "SELECT * FROM guru ORDER BY nama"
        cursor.execute(sql)
        guruAll = cursor.fetchall()

        # Slice data yang akan ditampilkan sebanyak 10 baris
        guru = guruAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = guruAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        if request.method == 'POST':
            # Jika user melakukan searching
            if request.form['search']:
                # Searching dengan mencocokkan nama guru
                try:
                    pencarian = request.form['search']
                    sql = "SELECT * FROM guru WHERE nama LIKE '%" + request.form['search'] + "%' ORDER BY nama"
                    cursor.execute(sql)
                    guru = cursor.fetchall()
                    lanjut = False
                except Exception as err:
                    Stat = False
                    flash('Terjadi kesalahan')
                    flash(err)
        # Jika databasenya kosong
        if not guru:
            guru = False
        closeDb()
        return render_template('admin/data_guru.html', count=counter, guru=guru, prev=prev, next=next, lanjut=lanjut, status=Stat, pencarian=pencarian)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_guru/<id>', methods = ['GET', 'POST'])
def editguru(id):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        cursor.execute('SELECT * FROM guru WHERE nuptk = %s', (id))
        data = cursor.fetchone()

        # Ambil list agama untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT * FROM agama')
        agam = cursor.fetchall()

        # Jika user menekan tombol submit, maka mengumpulkan semua data yang ada di form
        if request.method == 'POST':
            try:
                nuptk = request.form['nuptk']
                nama = request.form['nama']
                # Username = nama tanpa spasi ataupun simbol, hanya alphanumeric
                username = ''.join(us for us in nama if us.isalnum()).lower()
                jk = request.form['gender']
                agama = request.form['aga']
                email = request.form['email']
                telp = request.form['tel']
                detail = (nama, jk, agama, email, telp, nuptk)
                detail2 = (username, email, nuptk)
                # Update data guru
                cursor.execute(
                    'UPDATE guru SET nama = %s, jeniskelamin = %s, agama = %s, email = %s, telepon = %s WHERE nuptk = %s',
                    detail)
                conn.commit()
                # Update data user guru
                cursor.execute('UPDATE user SET username = %s, email = %s WHERE userid = %s', detail2)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('dataguru'))
        closeDb()
        return render_template('admin/edit_guru.html', data=data, ag=agam, gend=gend)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/hapus_guru/<id>', methods = ['GET', 'POST'])
def hapusguru(id):
    if verifLogin(1):
        openDb()
        global Stat
        cursor.execute('SELECT * FROM guru WHERE nuptk = %s', (id))
        data = cursor.fetchone()

        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                # Hapus guru dari daftar guru
                cursor.execute('DELETE FROM guru WHERE nuptk = %s', (id))
                conn.commit()
                # Hapus guru dari daftar user
                cursor.execute('DELETE FROM user WHERE userid = %s', (id))
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('dataguru'))
        closeDb()
        return render_template('admin/hapus_guru.html', data=data, gend=gend)
    else:
        return redirect(url_for('index'))
@app.route('/dashboard_admin/tambah_guru', methods = ['GET', 'POST'])
def tambahguru():
    if verifLogin(1):
        openDb()
        # Untuk masuk ke form upload guru
        tipe="guru"

        # Ambil list agama untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT * FROM agama')
        aga = cursor.fetchall()

        # Jika user mensubmit form
        if request.method == 'POST':
            try:
                nuptk = request.form['nuptk']
                nama = request.form['nama']
                # Username = nama tanpa spasi ataupun simbol, hanya alphanumeric
                username = ''.join(nama.split()).lower()
                jk = request.form.get('gender')
                ag = request.form.get('agama')
                em = request.form['email']
                tele = request.form['tel']
                detail = (nuptk, nama, jk, ag, em, tele)
                detail2 = (nuptk, username, em, tele, 2)
                # Tambahkan data ke table guru
                cursor.execute('INSERT INTO guru VALUES (%s, %s, %s, %s, %s, %s)', detail)
                conn.commit()

                # Tambahkan data ke table user sebagai guru
                cursor.execute('INSERT INTO user VALUES (%s, %s, %s, %s, %s)', detail2)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('dataguru'))
        closeDb()
        return render_template('admin/tambah_guru.html', tipe=tipe, aga=aga)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_siswa', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_admin/data_siswa/<int:page>', methods = ['GET', 'POST'])
def datasiswa(page):
    if verifLogin(1):
        openDb()
        global Stat
        lanjut = True
        prev = page - 1
        next = page + 1
        counter = prev*10
        # Ambil daftar siswa dari database
        sql = "SELECT s.nis, s.nama, s.jeniskelamin, s.agama, s.email, s.telepon, k.namakelas FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota LEFT JOIN kelas k ON r.kelas = k.kelas ORDER BY s.nama"
        cursor.execute(sql)
        siswaAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        siswa = siswaAll[counter:page*10]
        # Data slice halaman selanjutnya
        nextPage = siswaAll[page*10:next*10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        if request.method == 'POST':
            # Jika user melakukan searching
            if request.form['search']:
                # Searching dengan mencocokkan nama siswa
                try:
                    sql = "SELECT s.nis, s.nama, s.jeniskelamin, s.agama, s.email, s.telepon, k.namakelas FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota LEFT JOIN kelas k ON r.kelas = k.kelas WHERE s.nama LIKE '%" + request.form['search'] + "%' ORDER BY s.nama"
                    cursor.execute(sql)
                    siswa = cursor.fetchall()
                    lanjut = False
                except Exception as err:
                    Stat = False
                    flash('Terjadi kesalahan')
                    flash(err)
        # Jika databasenya kosong
        if not siswa:
            siswa = False
        return render_template('admin/data_siswa.html', count=counter, siswa=siswa, prev=prev, next=next, lanjut=lanjut, status=Stat)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_siswa/<id>', methods = ['GET', 'POST'])
def editsiswa(id):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil data siswa yang akan diedit
        sql = "SELECT s.nis, s.nama, s.jeniskelamin, s.agama, s.email, s.telepon, k.kelas, k.namakelas FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota LEFT JOIN kelas k ON r.kelas = k.kelas WHERE s.nis = %s"
        cursor.execute(sql, id)
        data = cursor.fetchone()

        # Ambil list agama untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT * FROM agama')
        agam = cursor.fetchall()

        # Ambil list kelas untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT kelas, namakelas FROM kelas')
        kls = cursor.fetchall()

        # Jika user menekan tombol submit, maka mengumpulkan semua data yang ada di form
        if request.method == 'POST':
            try:
                nis = request.form['nis']
                nama = request.form['nama']
                # Username = nama tanpa spasi ataupun simbol, hanya alphanumeric
                username = ''.join(us for us in nama if us.isalnum()).lower()
                jk = request.form['gender']
                agama = request.form['aga']
                email = request.form['email']
                telp = request.form['tel']
                kelas = request.form.get('kls')
                detail = (nama, jk, agama, email, telp, nis)
                detail2 = (kelas, nis)
                detail3 = (username, email, nis)
                # Update data siswa
                cursor.execute(
                    'UPDATE siswa SET nama = %s, jeniskelamin = %s, agama = %s, email = %s, telepon = %s WHERE nis = %s',
                    detail)
                conn.commit()
                # Update data kelas siswa
                cursor.execute('UPDATE rombel SET kelas = %s WHERE anggota = %s', detail2)
                conn.commit()
                # Update data user siswa
                cursor.execute('UPDATE user SET username = %s, email = %s WHERE userid = %s', detail3)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datasiswa'))
        return render_template('admin/edit_siswa.html', data=data, ag=agam, kls=kls, gend=gend)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/hapus_siswa/<id>', methods = ['GET', 'POST'])
def hapussiswa(id):
    if verifLogin(1):
        openDb()
        global Stat
        # Ambil data siswa yang akan dihapus
        sql = "SELECT s.nis, s.nama, s.jeniskelamin, s.agama, s.email, s.telepon, k.kelas, k.namakelas FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota LEFT JOIN kelas k ON r.kelas = k.kelas WHERE s.nis = %s"
        cursor.execute(sql, id)
        data = cursor.fetchone()

        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                # Hapus siswa dari daftar siswa
                cursor.execute('DELETE FROM siswa WHERE nis = %s', (id))
                conn.commit()
                # Hapus siswa dari daftar rombel kelas
                cursor.execute('DELETE FROM rombel WHERE anggota = %s', (id))
                conn.commit()
                # Hapus siswa dari daftar user
                cursor.execute('DELETE FROM user WHERE userid = %s', (id))
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datasiswa'))
        return render_template('admin/hapus_siswa.html', data=data, gend=gend)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/tambah_siswa', methods = ['GET', 'POST'])
def tambahsiswa():
    if verifLogin(1):
        openDb()
        # Untuk masuk ke form upload siswa
        tipe="siswa"

        # Ambil list agama untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT * FROM agama')
        aga = cursor.fetchall()

        # Ambil list kelas untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT kelas, namakelas FROM kelas')
        kls = cursor.fetchall()

        # Jika user mensubmit form
        if request.method == 'POST':
            try:
                nis = request.form['nis']
                nama = request.form['nama']
                # Username = nama tanpa spasi ataupun simbol, hanya alphanumeric
                username = ''.join(us for us in nama if us.isalnum()).lower()
                jk = request.form.get('gender')
                ag = request.form.get('agama')
                em = request.form['email']
                tele = request.form['tel']
                kelas = request.form.get('kls')
                detail = (nis, nama, jk, ag, em, tele)
                detail2 = (nis, username, em, tele, 3)
                detail3 = (kelas, nis)
                # Tambahkan data ke table siswa
                cursor.execute('INSERT INTO siswa VALUES (%s, %s, %s, %s, %s, %s)', detail)
                conn.commit()

                # Tambahkan data ke table user sebagai siswa
                cursor.execute('INSERT INTO user VALUES (%s, %s, %s, %s, %s)', detail2)
                conn.commit()

                # Tambahkan data siswa ke table rombel
                cursor.execute('INSERT INTO rombel VALUES (%s, %s)', detail3)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datasiswa'))
        closeDb()
        return render_template('admin/tambah_siswa.html', tipe=tipe, aga=aga, kls=kls)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/upload_file/<tipe>', methods = ['GET', 'POST'])
def formupload(tipe):
    if verifLogin(1):
        openDb()
        # Klasifikasi tipe file yang diupload
        if tipe == 'siswa':
            siswa = True
            guru = False
        elif tipe == 'guru':
            guru = True
            siswa = False

        # Jika user menekan tombol upload guru
        if request.method == 'POST' and guru:
            listnama = []
            listuser = []
            f = request.files['fileGuru']
            filename = app.config['UPLOAD_FOLDER'] + '/' + \
                       secure_filename(f.filename)
            try:
                f.save(filename)
                wb = load_workbook(filename=filename)
                ws = wb['Sheet1']
                # Membaca data mulai dari row ke 2
                cleanrows = ws.iter_rows(min_row=2)
                for id, nm, jk, ag, em, tl in cleanrows:
                    nuptk = id.value
                    nama = nm.value
                    jeniskelamin = jk.value
                    agama = ag.value
                    email = em.value
                    telepon = tl.value
                    # Username = nama tanpa spasi ataupun simbol, hanya alphanumeric
                    username = ''.join(us for us in nama if us.isalnum()).lower()
                    # Menyimpan data ke dalam tuple
                    templist = (nuptk, nama, jeniskelamin, agama, email, telepon)
                    templist2 = (nuptk, username, email, telepon, 2)
                    # Menyimpan tuple setiap row ke dalam sebuah list
                    listnama.append(templist)
                    listuser.append(templist2)
                # Tambah ke tabel guru
                sql = 'INSERT INTO guru VALUES (%s, %s, %s, %s, %s, %s)'
                cursor.executemany(sql, listnama)
                conn.commit()

                # Tambah ke tabel user
                sql = 'INSERT INTO user VALUES (%s, %s, %s, %s, %s)'
                cursor.executemany(sql, listuser)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('dataguru'))
        # Jika user menekan tombol upload siswa
        elif request.method == 'POST' and siswa:
            try:

                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datasiswa'))
        closeDb()
        return render_template('admin/form_upload.html', siswa=siswa, guru=guru, tipe=tipe)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_kelas', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_admin/data_kelas/<int:page>', methods = ['GET', 'POST'])
def datakelas(page):
    if verifLogin(1):
        openDb()
        global Stat
        lanjut = True
        prev = page - 1
        next = page + 1
        counter = prev * 10
        # Ambil daftar kelas, nama wali kelas, dan jumlah siswa dari database
        sql = "SELECT k.kelas, k.namakelas, g.nama, COUNT(r.anggota) FROM kelas k LEFT JOIN guru g ON k.walikelas = g.nuptk LEFT JOIN rombel r ON k.kelas = r.kelas GROUP BY k.kelas ORDER BY k.kelas ASC"
        cursor.execute(sql)
        kelasAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        kelas = kelasAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = kelasAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        # Jika databasenya kosong
        if not kelas:
            kelas = False
        closeDb()
        return render_template('admin/data_kelas.html', count=counter, kelas=kelas, prev=prev, next=next, lanjut=lanjut, status=Stat)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_kelas/<kode>', methods = ['GET', 'POST'])
def editkelas(kode):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat

        # Ambil data yang akan diedit
        cursor.execute(
            'SELECT k.kelas, k.namakelas, g.nama FROM kelas k LEFT JOIN guru g ON k.walikelas = g.nuptk WHERE k.kelas = %s',
            (kode))
        data = cursor.fetchone()

        # Ambil list guru yang belum menjabat sebagai wali kelas untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT nuptk, nama FROM guru WHERE NOT EXISTS (SELECT walikelas FROM kelas WHERE guru.nuptk = kelas.walikelas)')
        wali = cursor.fetchall()

        # Konfirm edit kelas
        if request.method == 'POST':
            try:
                namakelas = request.form['kelas']
                kodekelas = namakelas.replace(" ", "_")
                namawali = request.form['wali']
                # Cek id wali yang dipilih
                cursor.execute('SELECT nuptk FROM guru WHERE nama = %s', namawali)
                kodeguru = cursor.fetchone()
                info = (kodekelas, namakelas, kodeguru[0], kode)
                # Update data kelas
                cursor.execute('UPDATE kelas SET kelas = %s, namakelas = %s, walikelas = %s WHERE kelas = %s', info)
                conn.commit()
                # Update data rombel
                info2 = (kodekelas, kode)
                cursor.execute('UPDATE rombel SET kelas = %s WHERE kelas = %s', info2)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datakelas'))
        return render_template('admin/edit_kelas.html', data=data, wali=wali)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/hapus_kelas/<kode>', methods=['GET', 'POST'])
def hapuskelas(kode):
    if verifLogin(1):
        openDb()
        global Stat
        cursor.execute('SELECT k.kelas, k.namakelas, g.nama FROM kelas k LEFT JOIN guru g ON k.walikelas = g.nuptk WHERE k.kelas = %s', (kode))
        kelas = cursor.fetchone()

        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                cursor.execute('DELETE FROM kelas WHERE kelas = %s', (kode))
                conn.commit()
                # Meskipun data kelas di hapus,
                # Data rombel tidak perlu dihapus, sehingga ketika ditambahkan kembali data anggota kelasnya masih ada
                '''cursor.execute('DELETE FROM rombel WHERE kelas = %s', (kode))
                conn.commit()'''
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datakelas'))
        return render_template('admin/hapus_kelas.html', data=kelas)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/tambah_kelas', methods = ['GET', 'POST'])
def tambahkelas():
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil list guru yang belum menjabat sebagai wali kelas untuk ditampilkan pada pilihan dropdown
        cursor.execute(
            'SELECT nuptk, nama FROM guru WHERE NOT EXISTS (SELECT walikelas FROM kelas WHERE guru.nuptk = kelas.walikelas)')
        wali = cursor.fetchall()
        if request.method == 'POST':
            try:
                namakelas = request.form['kelas']
                kodekelas = namakelas.replace(" ", "_")
                walikelas = request.form.get('wali')
                detail = (kodekelas, namakelas, walikelas)
                cursor.execute('INSERT INTO kelas VALUES (%s, %s, %s)', detail)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datakelas'))
        closeDb()
        return render_template('admin/tambah_kelas.html', wali=wali)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_rombel/<kode>', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_admin/data_rombel/<kode>/<int:page>', methods = ['GET', 'POST'])
def datarombel(kode, page):
    if verifLogin(1):
        openDb()
        global Stat
        lanjut = True
        prev = page - 1
        next = page + 1
        counter=prev*10
        # Ambil nama kelas dari kode kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kode)
        kelas = cursor.fetchone()
        # Ambil daftar siswa kelas tersebut dari database
        sql = "SELECT s.nis, s.nama FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota WHERE r.kelas = %s ORDER BY s.nama"
        cursor.execute(sql, kode)
        rombelAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        rombel = rombelAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = rombelAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        # Jika databasenya kosong
        if not rombel:
            rombel = False
        closeDb()
        return render_template('admin/data_rombel.html', kode=kode, kelas=kelas, count=counter, rombel=rombel, prev=prev, next=next, lanjut=lanjut, status=Stat)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_rombel/<id>', methods = ['GET', 'POST'])
def editrombel(id):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil data siswa yang akan diedit
        sql = "SELECT s.nis, s.nama, k.kelas, k.namakelas FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota LEFT JOIN kelas k ON r.kelas = k.kelas WHERE s.nis = %s"
        cursor.execute(sql, id)
        data = cursor.fetchone()

        # Ambil list kelas untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT kelas, namakelas FROM kelas')
        kls = cursor.fetchall()

        # Jika user menekan tombol submit, maka mengumpulkan semua data yang ada di form
        if request.method == 'POST':
            try:
                nis = request.form['nis']
                kelas = request.form.get('kls')
                detail = (kelas, nis)

                # Update data kelas siswa
                cursor.execute('UPDATE rombel SET kelas = %s WHERE anggota = %s', detail)
                conn.commit()

                Stat = True
                flash('Data berhasil diubah')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datarombel', kode=data[2]))
        return render_template('admin/edit_rombel.html', data=data, kls=kls)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_jadwal', methods = ['GET', 'POST'])
def datajadwal():
    if verifLogin(1):
        openDb()
        # Ambil daftar kelas dari database
        sql = "SELECT kelas, namakelas FROM kelas ORDER BY namakelas"
        cursor.execute(sql)
        listKelas = cursor.fetchall()
        closeDb()
        return render_template('admin/data_jadwal.html', kelas=listKelas)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/jadwal/<kode>', methods = ['GET', 'POST'], defaults={'hari':"Senin"})
@app.route('/dashboard_admin/jadwal/<kode>/<string:hari>', methods = ['GET', 'POST'])
def jadwal(kode, hari):
    if verifLogin(1):
        openDb()
        global Stat
        tambah = True
        # Ambil nama kelas dari kode kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kode)
        kelas = cursor.fetchone()

        # Ambil data jadwal menurut kelas yang dipilih
        sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute), j.mapel, p.namamapel, g.nama FROM jadwal j LEFT JOIN guru g ON j.pengajar = g.nuptk LEFT JOIN mapel p ON j.mapel = p.kodemapel WHERE j.kelas = %s AND j.hari = %s ORDER BY j.jam"
        detail=(kode, hari)
        cursor.execute(sql, detail)
        jadwal = cursor.fetchall()

        # Ambil daftar hari belajar
        sql = "SELECT * FROM haribelajar ORDER BY hari DESC"
        cursor.execute(sql)
        listHari = cursor.fetchall()

        # Jika jadwal kelas tersebut sudah penuh
        if len(jadwal) >= 8:
            tambah = False

        # Jika jadwal kelas tersebut kosong
        if not jadwal:
            jadwal = False

        closeDb()
        return render_template('admin/detail_jadwal.html', kode=kode, hari=hari, kelas=kelas, jadwal=jadwal, listHari=listHari, status=Stat, tambah=tambah)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/tambah_jadwal', methods = ['GET', 'POST'])
def tambahjadwal():
    if verifLogin(1):
        openDb()
        global Stat
        # Ambil daftar kelas
        sql = "SELECT kelas, namakelas FROM kelas ORDER BY namakelas DESC"
        cursor.execute(sql)
        kls = cursor.fetchall()

        # Ambil daftar hari belajar
        sql = "SELECT * FROM haribelajar ORDER BY hari DESC"
        cursor.execute(sql)
        days = cursor.fetchall()

        if request.method == 'POST':
            try:
                Stat = True
                kode = request.form.get('kelas')
                hari = request.form.get('hari')
                # Cek juga jika jadwal yang ingin ditambah sudah penuh
                sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute), j.mapel, p.namamapel, g.nama FROM jadwal j LEFT JOIN guru g ON j.pengajar = g.nuptk LEFT JOIN mapel p ON j.mapel = p.kodemapel WHERE j.kelas = %s AND j.hari = %s ORDER BY j.jam"
                detail = (kode, hari)
                cursor.execute(sql, detail)
                jadwal = cursor.fetchall()
                if len(jadwal) >= 8:
                    Stat = False
                    flash('Jadwal kelas ini sudah penuh dan tidak bisa ditambah')
                    return redirect(url_for('jadwal', kode=kode, hari=hari))
                return redirect(url_for('addjadwal', kode=kode, hari=hari))
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('tambahjadwal'))
        return render_template('admin/tambah_jadwal.html', kls=kls, days=days)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/jadwal_baru/<kode>/<hari>', methods = ['GET', 'POST'])
def addjadwal(kode, hari):
    if verifLogin(1):
        openDb()
        global Stat
        # Ambil detail kelas jadwal yang akan ditambah
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kode)
        kls = cursor.fetchone()

        # Ambil daftar jam belajar dari jadwal kelas bersangkutan yang belum ada mapelnya
        sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute) FROM jambelajar j WHERE NOT EXISTS (SELECT w.jam FROM jadwal w WHERE j.jam = w.jam AND kelas = %s AND hari = %s)"
        a = (kode, hari)
        cursor.execute(sql, a)
        jam = cursor.fetchall()

        # Ambil daftar mapel
        sql = "SELECT kodemapel, namamapel FROM mapel ORDER BY namamapel"
        cursor.execute(sql)
        mapel = cursor.fetchall()

        # Ambil daftar guru pengajar biasa yang bukan staf admin/perpus
        sql = "SELECT g.nuptk, g.nama FROM guru g LEFT JOIN user u ON g.nuptk = u.userid WHERE u.access = 2"
        cursor.execute(sql)
        guru = cursor.fetchall()

        if request.method == 'POST':
            try:
                jm = request.form.get('jam')
                mp = request.form.get('mapel')
                gr = request.form.get('guru')
                jdwl = (kode, hari, jm, mp, gr)
                sql = "INSERT INTO jadwal VALUES (%s, %s, %s, %s, %s)"
                cursor.execute(sql, jdwl)
                conn.commit()
                Stat = True
                flash('Jadwal berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('jadwal', kode=kode, hari=hari))
        return render_template('admin/jadwal_baru.html', kelas=kls, hari=hari, jam=jam, mapel=mapel, guru=guru)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_jadwal/<kode>/<hari>/<jam>', methods = ['GET', 'POST'])
def editjadwal(kode, hari, jam):
    if verifLogin(1):
        openDb()
        global Stat
        # Ambil detail kelas jadwal yang akan diedit
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kode)
        kls = cursor.fetchone()

        # Ambil detail jam berdasarkan jadwal yang akan diedit
        sql = "SELECT jam, DATE_ADD(jam, interval 45 minute) FROM jambelajar WHERE jam = %s"
        cursor.execute(sql, jam)
        infojam = cursor.fetchone()

        # Ambil detail jadwal yang akan diedit
        sql = "SELECT kelas, hari, jam, mapel, pengajar FROM jadwal WHERE kelas = %s AND hari = %s AND jam = %s"
        detail = (kode, hari, jam)
        cursor.execute(sql, detail)
        jdw = cursor.fetchone()

        # Ambil daftar mapel
        sql = "SELECT kodemapel, namamapel FROM mapel ORDER BY namamapel"
        cursor.execute(sql)
        listmapel = cursor.fetchall()

        # Ambil daftar guru pengajar biasa yang bukan staf admin/perpus
        sql = "SELECT g.nuptk, g.nama FROM guru g LEFT JOIN user u ON g.nuptk = u.userid WHERE u.access = 2"
        cursor.execute(sql)
        listguru = cursor.fetchall()

        if request.method == 'POST':
            try:
                mp = request.form.get('mapel')
                gr = request.form.get('guru')
                jdwl = (mp, gr, jam, kode, hari)
                sql = "UPDATE jadwal SET mapel = %s, pengajar = %s WHERE jam = %s AND kelas = %s AND hari = %s"
                cursor.execute(sql, jdwl)
                conn.commit()
                Stat = True
                flash('Jadwal berhasil diubah')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return (redirect(url_for('jadwal', kode=kode, hari=hari)))
        return render_template('admin/edit_jadwal.html', data=jdw, kelas=kls, hari=hari, listmapel=listmapel, listguru=listguru, jam=infojam)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/hapus_jadwal/<kode>/<hari>/<jam>', methods = ['GET', 'POST'])
def hapusjadwal(kode, hari, jam):
    if verifLogin(1):
        openDb()
        global Stat
        # Ambil detail kelas jadwal yang akan dihapus
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kode)
        kls = cursor.fetchone()

        # Ambil detail jadwal yang akan dihapus
        sql = "SELECT j.kelas, j.hari, j.jam, DATE_ADD(j.jam, interval 45 minute), p.namamapel, g.nama FROM jadwal j LEFT JOIN mapel p ON j.mapel = p.kodemapel LEFT JOIN guru g ON j.pengajar = g.nuptk WHERE j.kelas = %s AND j.hari = %s AND j.jam = %s"
        detail = (kode, hari, jam)
        cursor.execute(sql, detail)
        jdw = cursor.fetchone()

        if request.method == 'POST':
            try:
                jdwl = (jam, kode, hari)
                sql = "DELETE FROM jadwal WHERE jam = %s AND kelas = %s AND hari = %s"
                cursor.execute(sql, jdwl)
                conn.commit()
                Stat = True
                flash('Jadwal berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return (redirect(url_for('jadwal', kode=kode, hari=hari)))
        return render_template('admin/hapus_jadwal.html', data=jdw, kelas=kls)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_mapel', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_admin/data_mapel/<int:page>', methods = ['GET', 'POST'])
def datamapel(page):
    if verifLogin(1):
        openDb()
        global Stat
        lanjut = True
        pencarian = ""
        prev = page - 1
        next = page + 1
        counter = prev*10
        # Ambil daftar mapel dari database
        sql = "SELECT * FROM mapel ORDER BY namamapel"
        cursor.execute(sql)
        mapelAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        mapel = mapelAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = mapelAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        if request.method == 'POST':
            # Jika user melakukan searching
            if request.form['search']:
                # Searching dengan nama atau kode mapel
                try:
                    pencarian = request.form['search']
                    sql = "SELECT * FROM mapel WHERE kodemapel LIKE '%" + \
                          request.form['search'] + "%' OR namamapel LIKE '%" + request.form[
                              'search'] + "%' ORDER BY namamapel"
                    cursor.execute(sql)
                    mapel = cursor.fetchall()
                    lanjut = False
                except Exception as err:
                    Stat = False
                    flash('Terjadi kesalahan')
                    flash(err)
        # Jika databasenya kosong
        if not mapel:
            mapel = False
        closeDb()
        return render_template('admin/data_mapel.html', count=counter, mapel=mapel, prev=prev, next=next, lanjut=lanjut, status=Stat, pencarian=pencarian)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_mapel/<kode>', methods = ['GET', 'POST'])
def editmapel(kode):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        cursor.execute('SELECT * FROM mapel WHERE kodemapel = %s', (kode))
        mapel = cursor.fetchone()

        # Jika user menekan tombol submit, maka mengumpulkan semua data yang ada di form
        if request.method == 'POST':
            try:
                nKode = request.form['kodemapel']
                nama = request.form['namamapel']
                detail = (nKode, nama, kode)
                print(detail)
                cursor.execute('UPDATE mapel SET kodemapel = %s, namamapel = %s WHERE kodemapel = %s', detail)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datamapel'))
        closeDb()
        return render_template('admin/edit_mapel.html', data=mapel)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/hapus_mapel/<kode>', methods = {'GET', 'POST'})
def hapusmapel(kode):
    if verifLogin(1):
        openDb()
        global Stat
        cursor.execute('SELECT * FROM mapel WHERE kodemapel = %s', (kode))
        mapel = cursor.fetchone()

        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                cursor.execute('DELETE FROM mapel WHERE kodemapel = %s', (kode))
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datamapel'))
        closeDb()
        return render_template('admin/hapus_mapel.html', data=mapel)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/tambah_mapel', methods = ['GET', 'POST'])
def tambahmapel():
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        if request.method == 'POST':
            try:
                kodemapel = request.form['kodemapel']
                namamapel = request.form['namamapel']
                detail = (kodemapel, namamapel)
                cursor.execute('INSERT INTO mapel VALUES (%s, %s)', detail)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datamapel'))
        return render_template('admin/tambah_mapel.html')
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/data_staf', methods = ['GET', 'POST'])
def datastaf():
    if verifLogin(1):
        openDb()
        # Ambil daftar staf dari database, left join dengan nama hak akses & nama guru
        sql = "SELECT u.userid, g.nama, u.email, h.accessname FROM user u LEFT JOIN hakakses h ON u.access = h.accessid LEFT JOIN guru g ON u.userid = g.nuptk WHERE u.username NOT IN('admin') AND u.access = 1 OR u.access = 4"
        cursor.execute(sql)
        staf = cursor.fetchall()
        # Jika data staf kosong
        if not staf:
            staf = False
        return render_template('admin/data_staf.html', staf=staf, status=Stat)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/edit_staf/<id>', methods = ['GET', 'POST'])
def editstaf(id):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat

        # Ambil data yang akan diedit
        cursor.execute('SELECT u.userid, g.nama, u.email FROM user u LEFT JOIN hakakses h ON u.access = h.accessid LEFT JOIN guru g ON u.userid = g.nuptk WHERE u.userid = %s', (id))
        data = cursor.fetchone()

        # Ambil list hak akses untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT accessname FROM hakakses WHERE accessid = 1 OR accessid = 4')
        role = cursor.fetchall()

        # Konfirm edit staf
        if request.method == 'POST':
            try:
                nuptk = request.form['nuptk']
                peran = request.form['role']
                # Cek id hak akses berdasarkan yang dipilih di form
                cursor.execute('SELECT accessid FROM hakakses WHERE accessname = %s', peran)
                rl = cursor.fetchone()
                info = (rl[0], nuptk)
                # Update data user dengan role staf yang dipilih
                cursor.execute('UPDATE user SET access = %s WHERE userid = %s', info)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datastaf'))
        closeDb()
        return render_template('admin/edit_staf.html', data=data, role=role)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/hapus_staf/<id>', methods = ['GET', 'POST'])
def hapusstaf(id):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil data dari staf yang dipilih untuk dihapus
        cursor.execute('SELECT u.userid, g.nama, u.email, h.accessname FROM user u LEFT JOIN hakakses h ON u.access = h.accessid LEFT JOIN guru g ON u.userid = g.nuptk WHERE u.userid = %s',(id))
        data = cursor.fetchone()

        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                # Di hapus dari staf, artinya hak akses dikembalikan menjadi guru biasa
                temp = (2, id)
                cursor.execute('UPDATE user SET access = %s WHERE userid = %s', temp)
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus dari staf')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('datastaf'))
        closeDb()
        return render_template('admin/hapus_staf.html', data=data)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/tambah_staf', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_admin/tambah_staf/<int:page>', methods = ['GET', 'POST'])
def tambahstaf(page):
    if verifLogin(1):
        openDb()
        global Stat, lanjut
        pencarian = ""
        prev = page - 1
        next = page + 1
        counter = prev * 10
        # Ambil daftar guru dari database yang bukan staf admin/perpus
        sql = "SELECT u.userid, g.nama, g.email FROM user u LEFT JOIN guru g ON u.userid = g.nuptk WHERE u.access = 2 ORDER BY nama"
        cursor.execute(sql)
        stafAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        staf = stafAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = stafAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        if request.method == 'POST':
            # Jika user melakukan searching
            if request.form['search']:
                # Searching dengan mencocokkan nama guru
                try:
                    pencarian = request.form['search']
                    sql = "SELECT u.userid, g.nama, g.email FROM user u LEFT JOIN guru g ON u.userid = g.nuptk WHERE u.access = 2 AND g.nama LIKE '%" + request.form[
                    'search'] + "%' ORDER BY nama"
                    cursor.execute(sql)
                    staf = cursor.fetchall()
                    lanjut = False
                except Exception as err:
                    Stat = False
                    flash('Terjadi kesalahan')
                    flash(err)
        # Jika databasenya kosong
        if not staf:
            staf = False
        closeDb()
        return render_template('admin/tambah_staf.html', count=counter, staf=staf, prev=prev, next=next, lanjut=lanjut, status=Stat, pencarian=pencarian)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_admin/konfirm_staf/<id>', methods = ['GET', 'POST'])
def addStaf(id):
    if verifLogin(1):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil list hak akses untuk ditampilkan pada pilihan dropdown
        cursor.execute('SELECT accessname FROM hakakses WHERE accessid = 1 OR accessid = 4')
        role = cursor.fetchall()

        # Ambil data yang akan ditambahkan
        cursor.execute(
            'SELECT u.userid, g.nama, u.email FROM user u LEFT JOIN hakakses h ON u.access = h.accessid LEFT JOIN guru g ON u.userid = g.nuptk WHERE u.userid = %s',
            (id))
        data = cursor.fetchone()

        if request.method == 'POST':
            try:
                nuptk = id
                namarole = request.form.get('role')
                cursor.execute('SELECT accessid FROM hakakses WHERE accessname = %s', namarole)
                rol = cursor.fetchone()
                info = (rol[0], nuptk)
                # Update data user dengan role staf yang dipilih
                cursor.execute('UPDATE user SET access = %s WHERE userid = %s', info)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('datastaf'))

        return render_template('admin/konfirm_staf.html', data=data, role=role)
    else:
        return redirect(url_for('index'))


''' Akses Guru '''
@app.route('/dashboard_guru', methods = ['GET', 'POST'])
def guru():
    if verifLogin(2):
        openDb()
        # Fetch hari & tanggal saat ini
        tanggal = date.today()
        d = tanggal.strftime("%d")
        m = cekBulan(tanggal.month)
        y = tanggal.strftime("%Y")
        # dt = tanggal
        dt = d + " " + m + " " + y
        dow = tanggal.isoweekday()
        # numHari = hari
        numHari = cekHari(dow)

        # Fetch data dari guru yang login
        sql = "SELECT nuptk, nama FROM guru WHERE nuptk = %s"
        cursor.execute(sql, session['id'])
        guru = cursor.fetchone()

        # Fetch data jadwal hari ini
        sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute), j.mapel, p.namamapel, k.namakelas FROM jadwal j LEFT JOIN kelas k ON j.kelas = k.kelas LEFT JOIN mapel p ON j.mapel = p.kodemapel WHERE j.pengajar = %s AND j.hari = %s ORDER BY j.jam"
        temp = (guru[0], numHari)
        cursor.execute(sql, temp)
        jadwal = cursor.fetchall()

        # Jika jadwal kelas tersebut kosong
        if not jadwal:
            jadwal = False;
        return render_template('guru/dashboard_guru.html', hari=numHari, tanggal=dt, jadwal=jadwal, guru=guru)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/cek_absensi', methods = ['GET', 'POST'])
def cekAbsensi():
    if verifLogin(2):
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)
        # Ambil daftar kelas yang diajar
        sql = "SELECT DISTINCT j.kelas, k.namakelas FROM jadwal j LEFT JOIN kelas k ON j.kelas = k.kelas WHERE j.pengajar = %s"
        cursor.execute(sql, session['id'])
        kls = cursor.fetchall()

        # Ambil daftar mapel yang diajar
        sql = "SELECT DISTINCT j.kelas, j.mapel, m.namamapel FROM jadwal j LEFT JOIN mapel m ON j.mapel = m.kodemapel WHERE j.pengajar = %s AND NOT EXISTS (SELECT m.kodemapel, m.namamapel FROM mapel m WHERE m.kodemapel = 'UPAC' AND m.kodemapel = j.mapel)"
        cursor.execute(sql, session['id'])
        mapel = cursor.fetchall()
        # Jika guru tidak memiliki kelas & mapel yang diajar
        if not kls:
            kls = False

        return render_template('guru/data_absensi.html', hari=numHari, kls=kls, mapel=mapel)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/data_absensi/<mapel>/<kelas>', methods = ['GET', 'POST'], defaults = {'page':1})
@app.route('/dashboard_guru/data_absensi/<mapel>/<kelas>/<int:page>', methods = ['GET', 'POST'])
def dataAbsensi(mapel, kelas, page):
    if verifLogin(2):
        openDb()
        global Stat, lanjut
        prev = page - 1
        next = page + 1
        counter = prev * 10
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)
        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()
        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()
        # Ambil data absensi yang ada
        sql = "SELECT DISTINCT WEEKDAY(tanggal), tanggal FROM presensi WHERE mapel = %s AND kelas = %s"
        mapelKelas = (matapel[0], kls[0])
        cursor.execute(sql, mapelKelas)
        absensiAll = cursor.fetchall()
        hariAbsen = []
        for x in range(1, 8):
            hariAbsen.append(cekHari(x))
        # Slice data yang akan ditampilkan sebanyak 10 baris
        absensi = absensiAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = absensiAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        # Jika belum ada data absensi
        if not absensi:
            absensi = False
        return render_template('guru/detail_absensi.html', count=counter, hari=numHari, map=matapel, kls=kls, prev=prev, next=next, absensi=absensi, lanjut=lanjut, hariAbsen=hariAbsen, status=Stat)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/info_absensi/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def infoAbsensi(mapel, kelas, tgl):
    if verifLogin(2):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil total kehadiran
        sql = "SELECT s.status, ifnull(status_count,0) as total FROM statuspresensi s LEFT JOIN (SELECT p.status, count(*) as status_count FROM presensi p WHERE mapel = %s AND kelas = %s AND tanggal = %s GROUP BY p.status) as total_status ON s.status = total_status.status ORDER BY FIELD(s.status,'Hadir','Izin','Sakit','Alpha');"
        mapelas = (mapel, kelas, tgl)
        cursor.execute(sql, mapelas)
        res = cursor.fetchall()

        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariAbsen = cekHari(nhari.isoweekday())

        # Ambil data absensi yang akan dipilih
        sql = "SELECT p.siswa, s.nama, p.status FROM presensi p LEFT JOIN siswa s ON p.siswa = s.nis WHERE p.mapel = %s AND p.kelas = %s AND p.tanggal = %s"
        cursor.execute(sql, mapelas)
        absensiKelas = cursor.fetchall()

        return render_template('guru/info_absensi.html', hari=numHari, absensi=absensiKelas, mapel=mapel, kelas=kelas, tgl=tgl, kls=kls, res=res, hariAbsen=hariAbsen)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/edit_absensi/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def editAbsensi(mapel, kelas, tgl):
    if verifLogin(2):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data absensi yang akan diedit
        sql = "SELECT p.siswa, s.nama, p.status FROM presensi p LEFT JOIN siswa s ON p.siswa = s.nis WHERE p.mapel = %s AND p.kelas = %s AND p.tanggal = %s"
        mapelas = (mapel, kelas, tgl)
        cursor.execute(sql, mapelas)
        absensiKelas = cursor.fetchall()

        # Ambil daftar opsi status presensi
        sql = "SELECT * FROM statuspresensi"
        cursor.execute(sql)
        statAbsensi = cursor.fetchall()

        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariAbsen = cekHari(nhari.isoweekday())

        if request.method == 'POST':
            try:
                jml = request.form['jumlahSiswa']
                for i in range(1, int(jml)+1):
                    nis = request.form['nis'+str(i)]
                    status = request.form['status'+str(i)]
                    temp = (status, nis, kelas, mapel, tgl)
                    sql = "UPDATE presensi SET status = %s WHERE siswa = %s AND kelas = %s AND mapel = %s AND tanggal = %s"
                    cursor.execute(sql, temp)
                    conn.commit()
                Stat = True
                flash('Data berhasil diubah')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
                closeDb()
            return redirect(url_for('dataAbsensi', mapel=mapel, kelas=kelas, page=1))

        return render_template('guru/edit_absensi.html', hari=numHari, absensi=absensiKelas, statAbsen=statAbsensi, mapel=mapel, kelas=kelas, tgl=tgl, kls=kls, hariAbsen=hariAbsen)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/hapus_absensi/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def hapusAbsensi(mapel, kelas, tgl):
    if verifLogin(2):
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()
        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()
        # Data hari/tanggal absen
        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariAbsen = cekHari(nhari.isoweekday())
        if request.method == 'POST':
            try:
                sql = "DELETE FROM presensi WHERE kelas = %s AND mapel = %s AND tanggal = %s"
                temp = (kelas, mapel, tgl)
                cursor.execute(sql, temp)
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('dataAbsensi', mapel=mapel, kelas=kelas, page=1))

        return render_template('guru/hapus_absensi.html', hari=numHari, kelas=kelas, mapel=mapel, tgl=tgl, kls=kls, matapel=matapel, hariAbsen=hariAbsen)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/absensi_baru', methods = ['GET', 'POST'])
def absensiBaru():
    if verifLogin(2):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil daftar kelas yang diajar
        sql = "SELECT DISTINCT j.kelas, k.namakelas FROM jadwal j LEFT JOIN kelas k ON j.kelas = k.kelas WHERE j.pengajar = %s"
        cursor.execute(sql, session['id'])
        kelas = cursor.fetchall()

        # Ambil daftar mapel yang diajar
        sql = "SELECT DISTINCT j.mapel, m.namamapel FROM jadwal j LEFT JOIN mapel m ON j.mapel = m.kodemapel WHERE j.pengajar = %s"
        cursor.execute(sql, session['id'])
        mapel = cursor.fetchall()

        if request.method == 'POST':
            try:
                kls = request.form.get('kelas')
                mpl = request.form.get('mapel')
                tgl = request.form.get('tanggal')
                sql = "SELECT DISTINCT hari FROM jadwal WHERE kelas = %s AND mapel = %s AND pengajar = %s"
                temp = (kls, mpl, session['id'])
                cursor.execute(sql, temp)
                hariAbsensi = cursor.fetchone()
                # Jika tanggal yang dipilih adalah hari sabtu atau minggu
                if datetime.strptime(tgl, '%Y-%m-%d').isoweekday() == 6 or datetime.strptime(tgl, '%Y-%m-%d').isoweekday() == 7:
                    Stat = False
                    flash('Tidak dapat menambah absensi di luar hari belajar')
                    return redirect(url_for('absensiBaru'))
                # Jika tanggal yang dipilih adalah hari di luar jadwal yang sudah ada
                elif datetime.strptime(tgl, '%Y-%m-%d').isoweekday() != cekHariR(hariAbsensi[0]):
                    Stat = False
                    flash('Tanggal tidak sesuai dengan jadwal mata pelajaran yang ada')
                    return redirect(url_for('absensiBaru'))
                closeDb()
                return redirect(url_for('tambahAbsensi', kelas=kls, mapel=mpl, tgl=tgl))
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
                closeDb()
            return redirect(url_for('cekAbsensi'))

        return render_template('guru/absensi_baru.html', hari=numHari, kelas=kelas, mapel=mapel)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/tambah_absensi/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def tambahAbsensi(kelas, mapel, tgl):
    if verifLogin(2):
        global Stat, siswa
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data murid di kelas tersebut
        sql = "SELECT r.anggota, s.nama FROM rombel r LEFT JOIN siswa s ON r.anggota = s.nis WHERE r.kelas = %s"
        cursor.execute(sql, kelas)
        siswa = cursor.fetchall()
        if request.method == 'POST':
            try:
                listTemp = []
                jml = request.form['jumlahSiswa']
                for i in range(1, int(jml)+1):
                    nis = request.form['nis'+str(i)]
                    status = request.form['status'+str(i)]
                    temp = (mapel, nis, kelas, tgl, status)
                    listTemp.append(temp)
                sql = "INSERT INTO presensi VALUES (%s, %s, %s, %s, %s)"
                cursor.executemany(sql, listTemp)
                conn.commit()

                Stat = True
                flash('Data berhasil ditambahkan')
                closeDb()
                return redirect(url_for('dataAbsensi', mapel=mapel, kelas=kelas, page=1))
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
                closeDb()
            return redirect(url_for('cekAbsensi'))

        return render_template('guru/tambah_absensi.html', hari=numHari, kelas=kls, mapel=mapel, tgl=tgl, siswa=siswa)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/cek_nilai/<jenis>', methods = ['GET', 'POST'])
def cekNilai(jenis):
    if verifLogin(2):
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil daftar kelas yang diajar
        sql = "SELECT DISTINCT j.kelas, k.namakelas FROM jadwal j LEFT JOIN kelas k ON j.kelas = k.kelas WHERE j.pengajar = %s"
        cursor.execute(sql, session['id'])
        kls = cursor.fetchall()

        # Ambil daftar mapel yang diajar kecuali upacara
        sql = "SELECT DISTINCT j.kelas, j.mapel, m.namamapel FROM jadwal j LEFT JOIN mapel m ON j.mapel = m.kodemapel WHERE j.pengajar = %s AND NOT EXISTS (SELECT m.kodemapel, m.namamapel FROM mapel m WHERE m.kodemapel = 'UPAC' AND m.kodemapel = j.mapel)"
        cursor.execute(sql, session['id'])
        mapel = cursor.fetchall()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        # Jika guru tidak memiliki kelas & mapel yang diajar
        if not kls:
            kls = False

        return render_template('guru/data_nilai.html', hari=numHari, kls=kls, mapel=mapel, jenis=jenis, jNilai=jNilai)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/data_nilai/<jenis>/<mapel>/<kelas>', methods = ['GET', 'POST'], defaults = {'page':1})
@app.route('/dashboard_guru/data_nilai/<jenis>/<mapel>/<kelas>/<int:page>', methods = ['GET', 'POST'])
def dataNilai(jenis, mapel, kelas, page):
    if verifLogin(2):
        openDb()
        global Stat, lanjut
        prev = page - 1
        next = page + 1
        counter = prev * 10
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil data daftar nilai yang dipilih
        sql = "SELECT DISTINCT WEEKDAY(tanggal), tanggal FROM nilaihitung WHERE mapel = %s AND kelas = %s AND nomor = %s"
        mapelKelas = (matapel[0], kls[0], jn)
        cursor.execute(sql, mapelKelas)
        nilaiAll = cursor.fetchall()
        hariNilai = []
        for x in range(1, 8):
            hariNilai.append(cekHari(x))
        # Slice data yang akan ditampilkan sebanyak 10 baris
        listNilai = nilaiAll[counter:page * 10]
        # Data slice halaman selanjutnya
        nextPage = nilaiAll[page * 10:next * 10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        # Jika belum ada data nilai
        if not listNilai:
            listNilai = False

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        return render_template('guru/detail_nilai.html', hari=numHari, kls=kls, map=matapel, listNilai=listNilai, hariNilai=hariNilai, jNilai=jNilai, jenis=jenis, lanjut=lanjut, status=Stat, count=counter, prev=prev, next=next)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/data_nilai_semester/<jenis>/<mapel>/<kelas>', methods = ['GET', 'POST'])
def nilaiSemester(jenis, mapel, kelas):
    if verifLogin(2):
        openDb()
        global Stat
        tempSum = 0
        tempNum = 0
        listSum = []
        avg = 0
        highest = 0
        lowest = 0
        tambah = False
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Cek jika sudah ada nilai di tabel db atau belum
        sql = "SELECT anggota FROM rombel WHERE kelas = %s"
        cursor.execute(sql, kelas)
        siswaKelas = cursor.fetchall()
        for x in siswaKelas:
            temp = "SELECT siswa FROM nilai WHERE mapel = %s AND siswa = %s AND kelas = %s"
            stk = (matapel[0], x, kls[0])
            cursor.execute(temp, stk)
            tempSiswa = cursor.fetchone()
            # Kalau belum ada/kosong, maka buat baru di tabel
            if not tempSiswa:
                temp2 = "INSERT INTO nilai (mapel, siswa, kelas) VALUES (%s, %s, %s)"
                cursor.execute(temp2, stk)
                conn.commit()

        # Ambil data nilai
        if jenis == "UTS":
            sql = "SELECT n.idnilai, n.siswa, s.nama, n.uts FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s"
        elif jenis == "UAS":
            sql = "SELECT n.idnilai, n.siswa, s.nama, n.uas FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s"
        mapelKelas = (matapel[0], kls[0])
        cursor.execute(sql, mapelKelas)
        listNilai = cursor.fetchall()

        # Jika belum ada data nilai
        for a in listNilai:
            if a[3] == None:
                tambah = True;
                listNilai = False
                break

        # Jika nilai tidak kosong, maka hitung rata-rata
        if listNilai:
            for b in listNilai:
                tempSum += b[3]
                listSum.append(b[3])
                tempNum += 1
            tempAvg = tempSum / tempNum
            # Untuk membatasi 2 digit angka desimal
            avg = f"{tempAvg:.2f}"
            highest = max(listSum)
            lowest = min(listSum)

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        return render_template('guru/nilai_semester.html', hari=numHari, kls=kls, map=matapel, listNilai=listNilai, tambah=tambah, jNilai=jNilai, jenis=jenis, status=Stat, avg=avg, max=highest, min=lowest)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/data_nilai_final/<mapel>/<kelas>', methods = ['GET', 'POST'])
def nilaiFinal(mapel, kelas):
    if verifLogin(2):
        openDb()
        global Stat, lengkap
        nilaiTugas = True
        nilaiPraktek = True
        nilaiUjian = True
        nilaiUTS = True
        nilaiUAS = True
        listNilai = []
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Cek jika sudah ada nilai di tabel db atau belum
        sql = "SELECT anggota FROM rombel WHERE kelas = %s  ORDER BY anggota ASC"
        cursor.execute(sql, kelas)
        siswaKelas = cursor.fetchall()
        for x in siswaKelas:
            temp = "SELECT siswa FROM nilai WHERE mapel = %s AND siswa = %s AND kelas = %s"
            stk = (matapel[0], x, kls[0])
            cursor.execute(temp, stk)
            tempSiswa = cursor.fetchone()
            # Kalau belum ada/kosong, maka buat baru di tabel
            if not tempSiswa:
                temp2 = "INSERT INTO nilai (mapel, siswa, kelas) VALUES (%s, %s, %s)"
                cursor.execute(temp2, stk)
                conn.commit()

        # Ambil semua nilai tugas
        sql = "SELECT DISTINCT tanggal FROM nilaihitung WHERE mapel = %s AND kelas = %s AND nomor = %s"
        temp = (matapel[0], kls[0], 1)
        cursor.execute(sql, temp)
        jmlTugas = cursor.fetchall()
        # Jika ada tugas, tabel tidak kosong
        if jmlTugas and siswaKelas:
            for x in siswaKelas:
                sql = "SELECT nilai FROM nilaihitung WHERE mapel = %s AND kelas = %s AND siswa = %s AND nomor = %s ORDER BY siswa ASC"
                temp = (matapel[0], kls[0], x, 1)
                cursor.execute(sql, temp)
                tempTugas = cursor.fetchall()
                tmpTgs = [a[0] for a in tempTugas]
                nilaiRata = sum(tmpTgs) / len(tmpTgs)
                sql = "UPDATE nilai SET tugas = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                tugas = (nilaiRata, x, kls[0], matapel[0])
                cursor.execute(sql, tugas)
                conn.commit()
        # Jika kosong
        else:
            nilaiTugas = False
            for x in siswaKelas:
                sql = "UPDATE nilai SET tugas = NULL WHERE siswa = %s AND kelas = %s AND mapel = %s"
                tglnull = (x, kls[0], matapel[0])
                cursor.execute(sql, tglnull)
                conn.commit()

        # Ambil semua nilai praktek
        sql = "SELECT DISTINCT tanggal FROM nilaihitung WHERE mapel = %s AND kelas = %s AND nomor = %s"
        temp = (matapel[0], kls[0], 2)
        cursor.execute(sql, temp)
        jmlPraktek = cursor.fetchall()
        # Jika ada praktek, tabel tidak kosong
        if jmlPraktek and siswaKelas:
            for x in siswaKelas:
                sql = "SELECT nilai FROM nilaihitung WHERE mapel = %s AND kelas = %s AND siswa = %s AND nomor = %s ORDER BY siswa ASC"
                temp = (matapel[0], kls[0], x, 2)
                cursor.execute(sql, temp)
                tempPraktek = cursor.fetchall()
                tmpPrtk = [a[0] for a in tempPraktek]
                nilaiRata = sum(tmpPrtk) / len(tmpPrtk)
                sql = "UPDATE nilai SET praktek = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                praktek = (nilaiRata, x, kls[0], matapel[0])
                cursor.execute(sql, praktek)
                conn.commit()
        else:
            nilaiPraktek = False
            for x in siswaKelas:
                sql = "UPDATE nilai SET praktek = NULL WHERE siswa = %s AND kelas = %s AND mapel = %s"
                prtknull = (x, kls[0], matapel[0])
                cursor.execute(sql, prtknull)
                conn.commit()

        # Ambil semua nilai ujian
        sql = "SELECT DISTINCT tanggal FROM nilaihitung WHERE mapel = %s AND kelas = %s AND nomor = %s"
        temp = (matapel[0], kls[0], 3)
        cursor.execute(sql, temp)
        jmlUjian = cursor.fetchall()
        # Jika ada ujian, tabel tidak kosong
        if jmlUjian and siswaKelas:
            for x in siswaKelas:
                sql = "SELECT nilai FROM nilaihitung WHERE mapel = %s AND kelas = %s AND siswa = %s AND nomor = %s ORDER BY siswa ASC"
                temp = (matapel[0], kls[0], x, 3)
                cursor.execute(sql, temp)
                tempUjian = cursor.fetchall()
                tmpUjn = [a[0] for a in tempUjian]
                nilaiRata = sum(tmpUjn) / len(tmpUjn)
                sql = "UPDATE nilai SET ujian = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                ujian = (nilaiRata, x, kls[0], matapel[0])
                cursor.execute(sql, ujian)
                conn.commit()
        else:
            nilaiUjian = False
            for x in siswaKelas:
                sql = "UPDATE nilai SET ujian = NULL WHERE siswa = %s AND kelas = %s AND mapel = %s"
                ujnnull = (x, kls[0], matapel[0])
                cursor.execute(sql, ujnnull)
                conn.commit()

        # Ambil nilai UTS
        sql = "SELECT uts FROM nilai WHERE mapel = %s AND kelas = %s ORDER BY siswa ASC"
        temp = (matapel[0], kls[0])
        cursor.execute(sql, temp)
        datanilaiUTS = cursor.fetchall()
        if not datanilaiUTS:
            nilaiUTS = False
            for x in siswaKelas:
                sql = "UPDATE nilai SET uts = NULL WHERE siswa = %s AND kelas = %s AND mapel = %s"
                utsnull = (x, kls[0], matapel[0])
                cursor.execute(sql, utsnull)
                conn.commit()

        # Ambil nilai UAS
        sql = "SELECT uas FROM nilai WHERE mapel = %s AND kelas = %s ORDER BY siswa ASC"
        temp = (matapel[0], kls[0])
        cursor.execute(sql, temp)
        datanilaiUAS = cursor.fetchall()
        if not datanilaiUAS:
            nilaiUAS = False
            for x in siswaKelas:
                sql = "UPDATE nilai SET uas = NULL WHERE siswa = %s AND kelas = %s AND mapel = %s"
                uasnull = (x, kls[0], matapel[0])
                cursor.execute(sql, uasnull)
                conn.commit()

        # Jika semua nilai lengkap, maka hitung nilai pengetahuan dan keterampilan
        if (nilaiTugas or nilaiPraktek) and nilaiUjian and nilaiUTS and nilaiUAS:
            lengkap = True
            sql = "SELECT tugas, praktek, ujian, uts, uas FROM nilai WHERE kelas = %s AND mapel = %s ORDER BY siswa ASC"
            temp = (kls[0], matapel[0])
            cursor.execute(sql, temp)
            nilaiSiswa = cursor.fetchall()
            x = 0
            for sw in siswaKelas:
                # Nilainya = 20% tugas atau praktek + 20% ujian + 30% uts + 30% uas
                # Jika tidak ada nilai praktek, maka samakan dengan tugas, dan sebaliknya
                if not nilaiTugas and nilaiPraktek:
                    keterampilan = 0.2 * float(nilaiSiswa[x][1]) + 0.2 * float(nilaiSiswa[x][2]) + 0.3 * float(
                        nilaiSiswa[x][3]) + 0.3 * float(nilaiSiswa[x][4])
                    pengetahuan = keterampilan
                elif not nilaiPraktek and nilaiTugas:
                    pengetahuan = 0.2 * float(nilaiSiswa[x][0]) + 0.2 * float(nilaiSiswa[x][2]) + 0.3 * float(
                        nilaiSiswa[x][3]) + 0.3 * float(nilaiSiswa[x][4])
                    keterampilan = pengetahuan
                else:
                    pengetahuan = 0.2 * float(nilaiSiswa[x][0]) + 0.2 * float(nilaiSiswa[x][2]) + 0.3 * float(nilaiSiswa[x][3]) + 0.3 * float(nilaiSiswa[x][4])
                    keterampilan = 0.2 * float(nilaiSiswa[x][1]) + 0.2 * float(nilaiSiswa[x][2]) + 0.3 * float(nilaiSiswa[x][3]) + 0.3 * float(nilaiSiswa[x][4])
                sql = "UPDATE nilai SET pengetahuan = %s, keterampilan = %s WHERE siswa = %s AND mapel = %s"
                temp = (pengetahuan, keterampilan, sw, matapel[0])
                cursor.execute(sql, temp)
                conn.commit()
                x += 1
            sql = "SELECT n.siswa, s.nama, n.tugas, n.praktek, n.ujian, n.uts, n.uas, n.pengetahuan, n.keterampilan FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.kelas = %s AND n.mapel = %s ORDER BY n.siswa ASC"
            temp = (kls[0], matapel[0])
            cursor.execute(sql, temp)
            listNilai = cursor.fetchall()
        else:
            lengkap = False

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        return render_template('guru/nilai_final.html', hari=numHari, kls=kls, map=matapel, lengkap=lengkap, jNilai=jNilai, status=Stat, listNilai=listNilai)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/info_nilai/<jenis>/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def infoNilai(jenis, mapel, kelas, tgl):
    if verifLogin(2):
        openDb()
        tempSum = 0
        tempNum = 0
        listSum = []
        avg = 0
        highest = 0
        lowest = 0
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariNilai = cekHari(nhari.isoweekday())

        # Ambil data nilai
        if jenis == "tugas":
            sql = "SELECT n.idnilai, n.siswa, s.nama, n.nilai FROM nilaihitung n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s AND n.nomor = 1 AND n.tanggal = %s"
        elif jenis == "praktek":
            sql = "SELECT n.idnilai, n.siswa, s.nama, n.nilai FROM nilaihitung n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s AND n.nomor = 2 AND n.tanggal = %s"
        elif jenis == "ujian":
            sql = "SELECT n.idnilai, n.siswa, s.nama, n.nilai FROM nilaihitung n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s AND n.nomor = 3 AND n.tanggal = %s"
        mapelKelas = (matapel[0], kls[0], tgl)
        cursor.execute(sql, mapelKelas)
        listNilai = cursor.fetchall()

        if listNilai:
            for b in listNilai:
                tempSum += b[3]
                listSum.append(b[3])
                tempNum += 1
            tempAvg = tempSum / tempNum
            # Untuk membatasi 2 digit angka desimal
            avg = f"{tempAvg:.2f}"
            highest = max(listSum)
            lowest = min(listSum)

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        return render_template('guru/info_nilai.html', hari=numHari, kls=kls, map=matapel, listNilai=listNilai, jNilai=jNilai, jenis=jenis, avg=avg, max=highest, min=lowest, hariNilai=hariNilai, tgl=tgl)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/nilai_baru/<jenis>/<mapel>/<kelas>', methods = ['GET', 'POST'])
def nilaiBaru(jenis, mapel, kelas):
    if verifLogin(2):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil daftar kelas yang diajar
        sql = "SELECT DISTINCT j.kelas, k.namakelas FROM jadwal j LEFT JOIN kelas k ON j.kelas = k.kelas WHERE j.pengajar = %s"
        cursor.execute(sql, session['id'])
        kls = cursor.fetchall()

        # Ambil daftar mapel yang diajar
        sql = "SELECT DISTINCT j.mapel, m.namamapel FROM jadwal j LEFT JOIN mapel m ON j.mapel = m.kodemapel WHERE j.pengajar = %s"
        cursor.execute(sql, session['id'])
        matapel = cursor.fetchall()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        if request.method == 'POST':
            try:
                jn = int(request.form.get('jenis'))
                cls = request.form.get('kelas')
                mpl = request.form.get('mapel')
                sql = "SELECT DISTINCT hari FROM jadwal WHERE kelas = %s AND mapel = %s AND pengajar = %s"
                temp = (cls, mpl, session['id'])
                cursor.execute(sql, temp)
                hariBelajar = cursor.fetchone()
                if jn == 1 or jn == 2 or jn == 3:
                    tgl = request.form.get('tanggal')
                    # Jika tanggal yang dipilih adalah hari sabtu atau minggu
                    if datetime.strptime(tgl, '%Y-%m-%d').isoweekday() == 6 or datetime.strptime(tgl, '%Y-%m-%d').isoweekday() == 7:
                        Stat = False
                        flash('Tidak dapat menambah nilai di luar hari belajar')
                        return redirect(url_for('nilaiBaru', jenis=jenis, kelas=kelas, mapel=mapel))
                    # Jika tanggal yang dipilih adalah hari di luar jadwal yang sudah ada
                    elif datetime.strptime(tgl, '%Y-%m-%d').isoweekday() != cekHariR(hariBelajar[0]):
                        Stat = False
                        flash('Tanggal tidak sesuai dengan jadwal mata pelajaran yang ada')
                        return redirect(url_for('nilaiBaru', jenis=jenis, kelas=kelas, mapel=mapel))
                    closeDb()
                    return redirect(url_for('tambahNilai', jenis=jenis, kelas=cls, mapel=mpl, tgl=tgl))
                elif jn == 4 or jn == 5:
                    return redirect(url_for('tambahNilaiSemester', jenis=jenis, kelas=cls, mapel=mpl))
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('cekNilai', jenis=jenis))

        return render_template('guru/nilai_baru.html', jn=jn, hari=numHari, kls=kls, map=matapel, kelas=kelas, mapel=mapel, jenis=jenis, jNilai=jNilai)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/tambah_nilai/<jenis>/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def tambahNilai(jenis, mapel, kelas, tgl):
    if verifLogin(2):
        global Stat
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariNilai = cekHari(nhari.isoweekday())

        # Ambil data murid di kelas tersebut
        sql = "SELECT r.anggota, s.nama FROM rombel r LEFT JOIN siswa s ON r.anggota = s.nis WHERE r.kelas = %s"
        cursor.execute(sql, kelas)
        siswa = cursor.fetchall()
        if request.method == 'POST':
            try:
                listTemp = []
                jml = request.form['jumlahSiswa']
                for i in range(1, int(jml) + 1):
                    nis = request.form['nis' + str(i)]
                    nilai = float(request.form['nilai' + str(i)])
                    temp = (matapel[0], nis, kls[0], jn, nilai, tgl)
                    listTemp.append(temp)
                sql = "INSERT INTO nilaihitung (mapel, siswa, kelas, nomor, nilai, tanggal) VALUES (%s, %s, %s, %s, %s, %s)"
                cursor.executemany(sql, listTemp)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('dataNilai', jenis=jenis, mapel=mapel, kelas=kelas, tgl=tgl, page=1))

        return render_template('guru/tambah_nilai.html', jn=jn, hari=numHari, kls=kls, map=matapel, tgl=tgl, siswa=siswa, hariNilai=hariNilai, jNilai=jNilai, jenis=jenis)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/edit_nilai/<jenis>/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def editNilai(jenis, mapel, kelas, tgl):
    if verifLogin(2):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil data nilai yang akan diedit
        sql = "SELECT n.siswa, s.nama, n.nilai FROM nilaihitung n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s AND n.nomor = %s AND n.tanggal = %s"
        mapelas = (mapel, kelas, jn, tgl)
        cursor.execute(sql, mapelas)
        listNilai = cursor.fetchall()

        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariNilai = cekHari(nhari.isoweekday())

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        if request.method == 'POST':
            try:
                jml = request.form['jumlahSiswa']
                for i in range(1, int(jml) + 1):
                    nis = request.form['nis' + str(i)]
                    nilai = float(request.form['nilai' + str(i)])
                    temp = (nilai, nis, kelas, mapel, jn, tgl)
                    sql = "UPDATE nilaihitung SET nilai = %s WHERE siswa = %s AND kelas = %s AND mapel = %s AND nomor = %s AND tanggal = %s"
                    cursor.execute(sql, temp)
                    conn.commit()
                Stat = True
                flash('Data berhasil diubah')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
                closeDb()
            return redirect(url_for('dataNilai', jenis=jenis, mapel=mapel, kelas=kelas, page=1))

        return render_template('guru/edit_nilai.html', jn=jn, hari=numHari, kls=kls, map=matapel, listNilai=listNilai, jNilai=jNilai, jenis=jenis, hariNilai=hariNilai, tgl=tgl)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/hapus_nilai/<jenis>/<mapel>/<kelas>/<tgl>', methods = ['GET', 'POST'])
def hapusNilai(jenis, mapel, kelas, tgl):
    if verifLogin(2):
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        # Data hari/tanggal nilai
        nhari = datetime.strptime(tgl, "%Y-%m-%d")
        hariNilai = cekHari(nhari.isoweekday())
        if request.method == 'POST':
            try:
                sql = "DELETE FROM nilaihitung WHERE nomor = %s AND kelas = %s AND mapel = %s AND tanggal = %s"
                temp = (jn, kls[0], matapel[0], tgl)
                cursor.execute(sql, temp)
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('dataNilai', jenis=jenis, mapel=mapel, kelas=kelas, page=1))

        return render_template('guru/hapus_nilai.html', jn=jn, hari=numHari, kls=kls, map=matapel, jNilai=jNilai, jenis=jenis, hariNilai=hariNilai, tgl=tgl)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/tambah_nilai_semester/<jenis>/<mapel>/<kelas>', methods = ['GET', 'POST'])
def tambahNilaiSemester(jenis, mapel, kelas):
    if verifLogin(2):
        global Stat
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        # Ambil data murid di kelas tersebut
        sql = "SELECT r.anggota, s.nama FROM rombel r LEFT JOIN siswa s ON r.anggota = s.nis WHERE r.kelas = %s"
        cursor.execute(sql, kelas)
        siswa = cursor.fetchall()
        if request.method == 'POST':
            try:
                jml = request.form['jumlahSiswa']
                for i in range(1, int(jml) + 1):
                    nis = request.form['nis' + str(i)]
                    nilai = float(request.form['nilai' + str(i)])
                    temp = (nilai, nis, kelas, mapel)
                    if jenis == "UTS":
                        sql = "UPDATE nilai SET uts = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                    elif jenis == "UAS":
                        sql = "UPDATE nilai SET uas = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                    cursor.execute(sql, temp)
                    conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('nilaiSemester', jenis=jenis, mapel=mapel, kelas=kelas))

        return render_template('guru/tambah_nilai.html', jn=jn, hari=numHari, kls=kls, map=matapel, siswa=siswa, jNilai=jNilai, jenis=jenis)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/edit_nilai_semester/<jenis>/<mapel>/<kelas>', methods = ['GET', 'POST'])
def editNilaiSemester(jenis, mapel, kelas):
    if verifLogin(2):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data nilai yang akan diedit
        if jenis == "UTS":
            sql = "SELECT n.siswa, s.nama, n.uts FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s"
        elif jenis == "UAS":
            sql = "SELECT n.siswa, s.nama, n.uas FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.kelas = %s"
        mapelas = (mapel, kelas)
        cursor.execute(sql, mapelas)
        listNilai = cursor.fetchall()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        if request.method == 'POST':
            try:
                jml = request.form['jumlahSiswa']
                for i in range(1, int(jml) + 1):
                    nis = request.form['nis' + str(i)]
                    nilai = float(request.form['nilai' + str(i)])
                    temp = (nilai, nis, kelas, mapel)
                    if jenis == "UTS":
                        sql = "UPDATE nilai SET uts = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                    elif jenis == "UAS":
                        sql = "UPDATE nilai SET uas = %s WHERE siswa = %s AND kelas = %s AND mapel = %s"
                    cursor.execute(sql, temp)
                    conn.commit()
                Stat = True
                flash('Data berhasil diubah')
                closeDb()
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
                closeDb()
            return redirect(url_for('nilaiSemester', jenis=jenis, mapel=mapel, kelas=kelas))

        return render_template('guru/edit_nilai.html', hari=numHari, kls=kls, map=matapel, listNilai=listNilai, jNilai=jNilai, jenis=jenis)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_guru/hapus_nilai_semester/<jenis>/<mapel>/<kelas>', methods = ['GET', 'POST'])
def hapusNilaiSemester(jenis, mapel, kelas):
    if verifLogin(2):
        openDb()
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil daftar siswa yang akan dihapus nilainya
        sql = "SELECT siswa FROM nilai WHERE kelas = %s AND mapel = %s"
        temp = (kelas, mapel)
        cursor.execute(sql, temp)
        listSiswa = cursor.fetchall()
        print(listSiswa[0])
        print(listSiswa[1])

        # Ambil jumlah nilai yang akan dipilih
        sql = "SELECT COUNT(*) FROM nilai WHERE kelas = %s AND mapel = %s"
        temp = (kelas, mapel)
        cursor.execute(sql, temp)
        jml = cursor.fetchone()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        if request.method == 'POST':
            try:
                for i in listSiswa:
                    temp = (kls[0], matapel[0], i[0])
                    if jenis == "UTS":
                        sql = "UPDATE nilai SET uts = NULL WHERE kelas = %s AND mapel = %s AND siswa = %s"
                    elif jenis == "UAS":
                        sql = "UPDATE nilai SET uas = NULL WHERE kelas = %s AND mapel = %s AND siswa = %s"
                    cursor.execute(sql, temp)
                    conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            closeDb()
            return redirect(url_for('nilaiSemester', jenis=jenis, mapel=mapel, kelas=kelas))

        return render_template('guru/hapus_nilai.html', jn=jn, hari=numHari, kls=kls, map=matapel, jNilai=jNilai, jenis=jenis)
    else:
        return redirect(url_for('index'))

'''
@app.route('/dashboard_guru/cetak_nilai/<kelas>', methods = ['GET', 'POST'])
def cetakNilai(kelas):
    if verifLogin(2):

        return render_template('guru/cetak_nilai.html')
    else:
        return redirect(url_for('index'))
'''

@app.route('/dashboard_guru/jadwal/<id>/<hari>', methods = ['GET', 'POST'])
def jadwalGuru(id, hari):
    if verifLogin(2):
        openDb()
        global Stat
        # Ambil data guru yang login
        sql = "SELECT nuptk, nama FROM guru WHERE nuptk = %s"
        cursor.execute(sql, id)
        guru = cursor.fetchone()

        # Ambil data jadwal menurut guru yang login
        sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute), j.mapel, p.namamapel, k.namakelas FROM jadwal j LEFT JOIN kelas k ON j.kelas = k.kelas LEFT JOIN mapel p ON j.mapel = p.kodemapel WHERE j.pengajar = %s AND j.hari = %s ORDER BY j.jam"
        detail = (id, hari)
        cursor.execute(sql, detail)
        jadwal = cursor.fetchall()

        # Ambil daftar hari belajar
        sql = "SELECT * FROM haribelajar ORDER BY hari DESC"
        cursor.execute(sql)
        listHari = cursor.fetchall()

        # Jika jadwal kelas tersebut kosong
        if not jadwal:
            jadwal = False;
        return render_template('guru/jadwal_guru.html', jadwal=jadwal, listHari=listHari, guru=guru, hari=hari)
    else:
        return redirect(url_for('index'))


''' Akses Siswa '''
@app.route('/dashboard_siswa', methods = ['GET', 'POST'])
def siswa():
    if verifLogin(3):
        openDb()
        # Fetch hari & tanggal saat ini
        tanggal = date.today()
        d = tanggal.strftime("%d")
        m = cekBulan(tanggal.month)
        y = tanggal.strftime("%Y")
        # dt = tanggal
        dt = d + " " + m + " " + y
        dow = tanggal.isoweekday()
        # numHari = hari
        numHari = cekHari(dow)

        # Fetch kelas dari siswa yang login
        sql = "SELECT r.kelas, k.namakelas from rombel r LEFT JOIN kelas k ON r.kelas = k.kelas WHERE r.anggota = %s"
        cursor.execute(sql, session['id'])
        kls = cursor.fetchone()

        # Fetch data jadwal hari ini
        sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute), j.mapel, p.namamapel, g.nama FROM jadwal j LEFT JOIN guru g ON j.pengajar = g.nuptk LEFT JOIN mapel p ON j.mapel = p.kodemapel WHERE j.kelas = %s AND j.hari = %s ORDER BY j.jam"
        temp = (kls[0], numHari)
        cursor.execute(sql, temp)
        jadwal = cursor.fetchall()

        # Jika jadwal kelas tersebut kosong
        if not jadwal:
            jadwal = False;
        return render_template('siswa/dashboard_siswa.html', hari=numHari, tanggal=dt, jadwal=jadwal, kelas=kls)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_siswa/cek_absen', methods = ['GET', 'POST'])
def cekAbsenSiswa():
    if verifLogin(3):
        openDb()
        # Fetch hari & tanggal saat ini
        tanggal = date.today()
        d = tanggal.strftime("%d")
        m = cekBulan(tanggal.month)
        y = tanggal.strftime("%Y")
        # dt = tanggal
        dt = d + " " + m + " " + y
        dow = tanggal.isoweekday()
        # numHari = hari
        numHari = cekHari(dow)

        # Ambil data kelas dari siswa yang login
        sql = "SELECT kelas FROM rombel WHERE anggota = %s"
        cursor.execute(sql, session['id'])
        kels = cursor.fetchone()

        # Ambil daftar mapel yang dipelajari selain upacara
        sql = "SELECT DISTINCT j.mapel, m.namamapel FROM jadwal j LEFT JOIN mapel m ON j.mapel = m.kodemapel WHERE j.kelas = %s AND NOT EXISTS (SELECT m.kodemapel, m.namamapel FROM mapel m WHERE m.kodemapel = 'UPAC' AND m.kodemapel = j.mapel)"
        cursor.execute(sql, kels)
        mapels = cursor.fetchall()

        closeDb()
        return render_template('siswa/absensi.html', hari=numHari, mapel=mapels, kelas=kels)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_siswa/absen/<mapel>', methods = ['GET', 'POST'])
def absenSiswa(mapel):
    if verifLogin(3):
        openDb()
        # Fetch hari & tanggal saat ini
        tanggal = date.today()
        d = tanggal.strftime("%d")
        m = cekBulan(tanggal.month)
        y = tanggal.strftime("%Y")
        # dt = tanggal
        dt = d + " " + m + " " + y
        dow = tanggal.isoweekday()
        # numHari = hari
        numHari = cekHari(dow)

        # Ambil data siswa yang sedang login
        sql = "SELECT nama FROM siswa WHERE nis = %s"
        cursor.execute(sql, session['id'])
        siswa = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kelas dari siswa tersebut
        sql = "SELECT kelas FROM rombel WHERE anggota = %s"
        cursor.execute(sql, session['id'])
        kelas = cursor.fetchone()

        # Ambil data absensi dari siswa yang bersangkutan
        sql = "SELECT WEEKDAY(tanggal), tanggal, status FROM presensi WHERE mapel = %s AND siswa = %s"
        det = (mapel, session['id'])
        cursor.execute(sql, det)
        abs = cursor.fetchall()

        if not abs:
            abs = False

        # Ambil total kehadiran
        sql = "SELECT s.status, ifnull(status_count,0) as total FROM statuspresensi s LEFT JOIN (SELECT p.status, count(*) as status_count FROM presensi p WHERE siswa = %s AND mapel = %s AND kelas = %s GROUP BY p.status) as total_status ON s.status = total_status.status ORDER BY FIELD(s.status,'Hadir','Izin','Sakit','Alpha');"
        temp = (session['id'], mapel, kelas)
        cursor.execute(sql, temp)
        res = cursor.fetchall()

        hariAbsen = []
        for x in range(1, 8):
            hariAbsen.append(cekHari(x))

        closeDb()
        return render_template('siswa/absensiSiswa.html', siswa=siswa, absensi=abs, hari=numHari, kelas=kelas, hariAbsen=hariAbsen, map=matapel, res=res)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_siswa/cek_nilai/<jenis>', methods = ['GET', 'POST'])
def cekNilaiSiswa(jenis):
    if verifLogin(3):
        openDb()
        global Stat
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data kelas
        sql = "SELECT r.kelas, k.namakelas FROM rombel r LEFT JOIN kelas k ON r.kelas = k.kelas WHERE r.anggota = %s"
        kelas = (session['id'])
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil daftar mapel yang dipelajari selain upacara
        sql = "SELECT DISTINCT j.mapel, m.namamapel FROM jadwal j LEFT JOIN mapel m ON j.mapel = m.kodemapel WHERE j.kelas = %s AND NOT EXISTS (SELECT m.kodemapel, m.namamapel FROM mapel m WHERE m.kodemapel = 'UPAC' AND m.kodemapel = j.mapel)"
        cursor.execute(sql, kls[0])
        mapels = cursor.fetchall()

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        return render_template('siswa/cekNilaiSiswa.html', jenis=jenis, kls=kls, hari=numHari, mapel=mapels, jNilai=jNilai)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_siswa/nilai/<jenis>/<mapel>', methods = ['GET', 'POST'])
def nilaiSiswa(jenis, mapel):
    if verifLogin(3):
        openDb()
        global Stat, tipe
        tanggal = date.today()
        dow = tanggal.isoweekday()
        numHari = cekHari(dow)

        # Ambil data siswa yang sedang login
        sql = "SELECT nama FROM siswa WHERE nis = %s"
        cursor.execute(sql, session['id'])
        siswa = cursor.fetchone()

        # Ambil data kelas
        sql = "SELECT r.kelas, k.namakelas FROM rombel r LEFT JOIN kelas k ON r.kelas = k.kelas WHERE r.anggota = %s"
        kelas = (session['id'])
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data mapel
        sql = "SELECT kodemapel, namamapel FROM mapel WHERE kodemapel = %s"
        cursor.execute(sql, mapel)
        matapel = cursor.fetchone()

        # Ambil data kodenilai berdasarkan jenis nilai yang diakses
        sql = "SELECT kodenilai FROM jenisnilai WHERE namanilai = %s"
        cursor.execute(sql, jenis)
        jn = cursor.fetchone()

        # Ambil data nilai yang dipilih
        avg = False
        if jn[0] == 1 or jn[0] == 2 or jn[0] == 3:
            tipe = 1
            sql = "SELECT WEEKDAY(tanggal), tanggal, nilai FROM nilaihitung WHERE mapel = %s AND siswa = %s AND nomor = %s"
            mapelas = (matapel[0], session['id'], jn[0])
            cursor.execute(sql, mapelas)
            dataNilai = cursor.fetchall()
            if dataNilai:
                tempSum = 0
                tempNum = 0
                listSum = []
                for a in dataNilai:
                    tempSum += a[2]
                    listSum.append(a[2])
                    tempNum += 1
                tempAvg = tempSum / tempNum
                # Untuk membatasi 2 digit angka desimal
                avg = f"{tempAvg:.2f}"
        elif jn[0] == 4 or jn[0] == 5:
            tipe = 2
            sql = "SELECT n.siswa, s.nama, n.uts, n.uas FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.siswa = %s"
            temp = (matapel[0], session['id'])
            cursor.execute(sql, temp)
            dataNilai = cursor.fetchall()
        elif jn[0] == 6:
            tipe = 3
            sql = "SELECT n.siswa, s.nama, n.pengetahuan, n.keterampilan FROM nilai n LEFT JOIN siswa s ON n.siswa = s.nis WHERE n.mapel = %s AND n.siswa = %s"
            temp = (matapel[0], session['id'])
            cursor.execute(sql, temp)
            dataNilai = cursor.fetchall()
        else:
            dataNilai = []

        hariNilai = []
        for x in range(1, 8):
            hariNilai.append(cekHari(x))

        if not dataNilai:
            dataNilai = False

        # Ambil daftar jenis nilai
        sql = "SELECT namanilai from jenisnilai ORDER BY kodenilai ASC"
        cursor.execute(sql)
        temp = cursor.fetchall()
        jNilai = []
        for x in temp:
            jNilai.append(x[0])

        return render_template('siswa/nilaiSiswa.html', siswa=siswa, jenis=jenis, hariNilai=hariNilai, kls=kls, map=matapel, hari=numHari, jNilai=jNilai, dataNilai=dataNilai, tipe=tipe, avg=avg)
    else:
        return redirect(url_for('index'))
@app.route('/dashboard_siswa/jadwal/<kelas>/<hari>', methods = ['GET','POST'])
def jadwalSiswa(kelas, hari):
    if verifLogin(3):
        openDb()
        global Stat
        # Ambil nama kelas dari kode kelas
        sql = "SELECT kelas, namakelas FROM kelas WHERE kelas = %s"
        cursor.execute(sql, kelas)
        kls = cursor.fetchone()

        # Ambil data jadwal menurut kelas yang dipilih
        sql = "SELECT j.jam, DATE_ADD(j.jam, interval 45 minute), j.mapel, p.namamapel, g.nama FROM jadwal j LEFT JOIN guru g ON j.pengajar = g.nuptk LEFT JOIN mapel p ON j.mapel = p.kodemapel WHERE j.kelas = %s AND j.hari = %s ORDER BY j.jam"
        detail=(kelas, hari)
        cursor.execute(sql, detail)
        jadwal = cursor.fetchall()

        # Ambil daftar hari belajar
        sql = "SELECT * FROM haribelajar ORDER BY hari DESC"
        cursor.execute(sql)
        listHari = cursor.fetchall()

        # Jika jadwal kelas tersebut kosong
        if not jadwal:
            jadwal = False;

        closeDb()
        return render_template('siswa/jadwal_siswa.html', kls=kls, hari=hari, jadwal=jadwal, listHari=listHari)
    else:
        return redirect(url_for('index'))


''' Akses Perpustakaan '''
@app.route('/dashboard_perpus', methods = ['GET', 'POST'])
def perpus():
    if verifLogin(4):
        openDb()
        # Fetch hari & tanggal saat ini
        tanggal = date.today()
        d = tanggal.strftime("%d")
        m = cekBulan(tanggal.month)
        y = tanggal.strftime("%Y")
        # dt = tanggal
        dt = d + " " + m + " " + y
        dow = tanggal.isoweekday()
        # numHari = hari
        numHari = cekHari(dow)

        # Fetch jumlah buku
        sql = "SELECT COUNT(*) FROM databuku"
        cursor.execute(sql)
        res = cursor.fetchone()
        numBuku = res[0]

        # Fetch jumlah total peminjaman
        sql = "SELECT COUNT(*) FROM peminjamanbuku"
        cursor.execute(sql)
        res = cursor.fetchone()
        totPeminjaman = res[0]

        # Fetch jumlah buku yang sedang dipinjam
        sql = "SELECT COUNT(*) FROM peminjamanbuku WHERE statuspinjam = 'Dipinjam'"
        cursor.execute(sql)
        res = cursor.fetchone()
        ongoingPeminjaman = res[0]
        closeDb()
        return render_template('perpus/dashboard_perpus.html', hari=numHari, tanggal=dt, buku=numBuku, total=totPeminjaman, pinjam=ongoingPeminjaman)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/buku', methods = ['GET', 'POST'], defaults={'page':1})
@app.route('/dashboard_perpus/buku/<int:page>', methods = ['GET', 'POST'])
def buku(page):
    if verifLogin(4):
        openDb()
        global lanjut, Stat
        pencarian = ""
        prev = page - 1
        next = page + 1
        counter = prev*10
        # Ambil daftar buku dari database
        sql = "SELECT isbn, namabuku, namapenulis, namapenerbit, tahunterbit FROM databuku ORDER BY namabuku"
        cursor.execute(sql)
        bukuAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        buku = bukuAll[counter:page*10]
        # Data slice halaman selanjutnya
        nextPage = bukuAll[page*10:next*10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        if request.method == 'POST':
            # Jika user melakukan searching
            if request.form['search']:
                # Searching dengan mencocokkan nama buku
                try:
                    pencarian = request.form['search']
                    sql = "SELECT isbn, namabuku, namapenulis, namapenerbit, tahunterbit FROM databuku WHERE isbn LIKE '%" + request.form['search'] + "%' OR namabuku LIKE '%" + request.form['search'] + "%' OR namapenulis LIKE '%" + request.form['search'] + "%' OR namapenerbit LIKE '%" + request.form['search'] + "%' OR tahunterbit LIKE '%" + request.form['search'] + "%' ORDER BY namabuku"
                    cursor.execute(sql)
                    buku = cursor.fetchall()
                    lanjut = False
                except Exception as err:
                    Stat = False
                    flash('Terjadi kesalahan')
                    flash(err)
        # Jika databasenya kosong
        if not buku:
            buku = False
        closeDb()
        return render_template('perpus/data_buku.html', count=counter, buku=buku, prev=prev, next=next, lanjut=lanjut, status=Stat, pencarian=pencarian)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/edit_buku/<id>', methods = ['GET', 'POST'])
def editbuku(id):
    if verifLogin(4):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil data buku yang akan diedit
        sql = "SELECT isbn, namabuku, namapenulis, namapenerbit, tahunterbit FROM databuku WHERE isbn = %s"
        cursor.execute(sql, id)
        data = cursor.fetchone()
        # Jika user menekan tombol submit, maka mengumpulkan semua data yang ada di form untuk diedit
        if request.method == 'POST':
            try:
                isbn = request.form['isbn']
                judul = request.form['judul']
                penulis = request.form['penulis']
                penerbit = request.form['penerbit']
                tahun = request.form['tahunterbit']
                detail = (isbn, judul, penulis, penerbit, tahun, id)
                # Update data buku
                cursor.execute(
                    'UPDATE databuku SET isbn = %s, namabuku = %s, namapenulis = %s, namapenerbit = %s, tahunterbit = %s WHERE isbn = %s',
                    detail)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('buku'))
        closeDb()
        return render_template('perpus/edit_buku.html', data=data)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/hapus_buku/<id>', methods = ['GET', 'POST'])
def hapusbuku(id):
    if verifLogin(4):
        openDb()
        global Stat
        # Ambil data buku yang akan dihapus
        cursor.execute('SELECT isbn, namabuku, namapenulis, namapenerbit, tahunterbit FROM databuku WHERE isbn = %s', (id))
        data = cursor.fetchone()
        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                cursor.execute('DELETE FROM databuku WHERE isbn = %s', (id))
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('buku'))
        closeDb()
        return render_template('perpus/hapus_buku.html', data=data)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/tambah_buku', methods = ['GET', 'POST'])
def tambahbuku():
    if verifLogin(4):
        openDb()
        # Status berhasil & error
        global Stat
        # Jika user mengkonfirm tambah data buku
        if request.method == 'POST':
            try:
                isbn = request.form['isbn']
                judul = request.form['judulbuku']
                penulis = request.form['penulis']
                penerbit = request.form['penerbit']
                tahun = request.form['tahunterbit']
                detail = (isbn, judul, penulis, penerbit, tahun)
                cursor.execute('INSERT INTO databuku VALUES (%s, %s, %s, %s, %s)', detail)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('buku'))
        closeDb()
        return render_template('perpus/tambah_buku.html')
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/peminjaman', methods = ['GET', 'POST'], defaults={'page':1, 'status':"Dipinjam"})
@app.route('/dashboard_perpus/peminjaman/<string:status>/<int:page>', methods = ['GET', 'POST'])
def peminjaman(status, page):
    if verifLogin(4):
        openDb()
        global lanjut, Stat
        listStatus = (['Dipinjam'], ['Dikembalikan'])
        pencarian = ""
        prev = page - 1
        next = page + 1
        counter = prev*10
        # Ambil daftar peminjaman dari database
        sql = "SELECT p.idpeminjaman, s.nama, k.namakelas, b.namabuku, p.tanggalpinjam, p.statuspinjam, p.tanggalpengembalian FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis WHERE p.statuspinjam = 'Dipinjam' ORDER BY p.idpeminjaman"
        if status == "Dikembalikan":
            sql = "SELECT p.idpeminjaman, s.nama, k.namakelas, b.namabuku, p.tanggalpinjam, p.statuspinjam, p.tanggalpengembalian FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis WHERE p.statuspinjam = 'Dikembalikan' OR p.statuspinjam = 'Rusak' OR p.statuspinjam = 'Hilang' ORDER BY p.idpeminjaman"
        cursor.execute(sql)
        peminjamanAll = cursor.fetchall()
        # Slice data yang akan ditampilkan sebanyak 10 baris
        peminjaman = peminjamanAll[counter:page*10]
        # Data slice halaman selanjutnya
        nextPage = peminjamanAll[page*10:next*10]
        # Jika halaman selanjutnya kosong
        if not nextPage:
            lanjut = False
        if request.method == 'POST':
            # Jika user melakukan searching
            if request.form['search']:
                try:
                    pencarian = request.form['search']
                    sql = "SELECT p.idpeminjaman, s.nama, k.namakelas, b.namabuku, p.tanggalpinjam, p.statuspinjam, p.tanggalpengembalian FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis WHERE s.nama LIKE '%" + request.form['search'] + "%' OR k.namakelas LIKE '%" + request.form['search'] + "%' OR b.namabuku LIKE '%" + request.form['search'] + "%' OR p.statuspinjam LIKE '%" + request.form['search'] + "%' OR p.tanggalpinjam LIKE '%" + request.form['search'] + "%' AND p.statuspinjam = 'Dipinjam' ORDER BY p.idpeminjaman"
                    if status == 'Dikembalikan':
                        sql = "SELECT p.idpeminjaman, s.nama, k.namakelas, b.namabuku, p.tanggalpinjam, p.statuspinjam, p.tanggalpengembalian FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis WHERE s.nama LIKE '%" + \
                              request.form['search'] + "%' OR k.namakelas LIKE '%" + request.form[
                                  'search'] + "%' OR b.namabuku LIKE '%" + request.form[
                                  'search'] + "%' OR p.statuspinjam LIKE '%" + request.form[
                                  'search'] + "%' OR p.tanggalpinjam LIKE '%" + request.form[
                                  'search'] + "%' AND p.statuspinjam = 'Dikembalikan' OR p.statuspinjam = 'Hilang' OR p.statuspinjam = 'Rusak' ORDER BY p.idpeminjaman"
                    cursor.execute(sql)
                    peminjaman = cursor.fetchall()
                    lanjut = False
                except Exception as err:
                    Stat = False
                    flash('Terjadi kesalahan')
                    flash(err)
        # Jika databasenya kosong
        if not peminjaman:
            peminjaman = False
        closeDb()
        return render_template('perpus/data_peminjaman.html', statuspinjam=status, count=counter, peminjaman=peminjaman, prev=prev, next=next, lanjut=lanjut, status=Stat, pencarian=pencarian, listStatus=listStatus)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/edit_peminjaman/<id>', methods = ['GET', 'POST'])
def editpeminjaman(id):
    if verifLogin(4):
        openDb()
        global Stat
        listStatus = (['Dikembalikan'], ['Rusak'], ['Hilang'])
        # Ambil data peminjaman yang akan diedit
        sql = "SELECT p.idpeminjaman, s.nama, k.namakelas, b.namabuku, p.tanggalpinjam, p.statuspinjam, p.tanggalpengembalian FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis WHERE p.idpeminjaman = %s"
        cursor.execute(sql, id)
        peminjaman = cursor.fetchone()
        if request.method == 'POST':
            try:
                statuspinj = request.form['status']
                tanggalkemb = request.form['tanggalkembali']
                # Update data peminjaman
                info = (statuspinj, tanggalkemb, id)
                cursor.execute('UPDATE peminjamanbuku SET statuspinjam = %s, tanggalpengembalian = %s WHERE idpeminjaman = %s', info)
                conn.commit()
                Stat = True
                flash('Data berhasil diubah')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('peminjaman'))
        closeDb()
        return render_template('perpus/edit_peminjaman.html', id=id, data=peminjaman, listStatus=listStatus)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/hapus_peminjaman/<id>', methods = ['GET', 'POST'])
def hapuspeminjaman(id):
    if verifLogin(4):
        openDb()
        global Stat
        # Ambil data peminjaman yang akan dihapus
        sql = "SELECT p.idpeminjaman, s.nama, k.namakelas, b.namabuku, p.tanggalpinjam, p.statuspinjam, p.tanggalpengembalian FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis WHERE p.idpeminjaman = %s"
        cursor.execute(sql, id)
        data = cursor.fetchone()
        # Jika user mengkonfirm hapusdata
        if request.method == 'POST':
            try:
                cursor.execute('DELETE FROM peminjamanbuku WHERE idpeminjaman = %s', id)
                conn.commit()
                Stat = True
                flash('Data berhasil dihapus')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('peminjaman'))
        closeDb()
        return render_template('perpus/hapus_peminjaman.html', data=data)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/tambah_peminjaman', methods = ['GET', 'POST'])
def tambahpeminjaman():
    if verifLogin(4):
        openDb()
        # Status berhasil & error
        global Stat
        # Ambil daftar siswa
        sql = "SELECT s.nis, s.nama, k.namakelas FROM siswa s LEFT JOIN rombel r ON s.nis = r.anggota LEFT JOIN kelas k ON k.kelas = r.kelas ORDER BY s.nama"
        cursor.execute(sql)
        siswa = cursor.fetchall()
        # Ambil daftar buku
        sql = "SELECT isbn, namabuku FROM databuku ORDER BY namabuku"
        cursor.execute(sql)
        buku = cursor.fetchall()
        # Jika user mengkonfirm tambah data peminjaman
        if request.method == 'POST':
            try:
                peminjam = request.form.get('peminjam')
                judulbuku = request.form.get('buku')
                sql = "SELECT isbn FROM databuku WHERE namabuku = %s"
                cursor.execute(sql, judulbuku)
                buku = cursor.fetchone()

                cursor.execute('SELECT kelas from rombel WHERE anggota = %s', peminjam)
                kls = cursor.fetchone()
                tanggal = request.form['tanggalpinjam']
                detail = (peminjam, kls[0], buku[0], tanggal, 'Dipinjam')
                cursor.execute('INSERT INTO peminjamanbuku VALUES (NULL, %s, %s, %s, %s, %s, NULL)', detail)
                conn.commit()
                Stat = True
                flash('Data berhasil ditambahkan')
            except Exception as err:
                Stat = False
                flash('Terjadi kesalahan')
                flash(err)
            return redirect(url_for('peminjaman'))
        closeDb()
        return render_template('perpus/tambah_peminjaman.html', siswa=siswa, buku=buku)
    else:
        return redirect(url_for('index'))

@app.route('/dashboard_perpus/cetak_peminjaman', methods = ['GET', 'POST'])
def cetakpeminjaman():
    if verifLogin(4):
        openDb()
        global lanjut, Stat
        # Ambil daftar peminjaman dari database
        sql = "SELECT p.idpeminjaman, s.nama, b.isbn, p.tanggalpinjam FROM peminjamanbuku p LEFT JOIN kelas k ON p.kelas = k.kelas LEFT JOIN databuku b ON p.nomorbuku = b.isbn LEFT JOIN siswa s ON p.peminjam = s.nis ORDER BY p.idpeminjaman"
        cursor.execute(sql)
        peminjamanAll = cursor.fetchall()

        # Jika databasenya kosong
        if not peminjamanAll:
            peminjamanAll = False
        else:
            res = []
            sql = "SELECT count(idpeminjaman) FROM peminjamanbuku WHERE statuspinjam = 'Dipinjam'"
            cursor.execute(sql)
            pinjam = cursor.fetchall()
            res.append(pinjam[0])
            sql = "SELECT count(idpeminjaman) FROM peminjamanbuku WHERE statuspinjam = 'Dikembalikan'"
            cursor.execute(sql)
            kembali = cursor.fetchall()
            res.append(kembali[0])
            sql = "SELECT count(idpeminjaman) FROM peminjamanbuku WHERE statuspinjam = 'Rusak'"
            cursor.execute(sql)
            rsk = cursor.fetchall()
            res.append(rsk[0])
            sql = "SELECT count(idpeminjaman) FROM peminjamanbuku WHERE statuspinjam = 'Hilang'"
            cursor.execute(sql)
            hlg = cursor.fetchall()
            res.append(hlg[0])
        closeDb()
        return render_template('perpus/cetak_pinjam.html', peminjaman=peminjamanAll, res=res)
    else:
        return redirect(url_for('index'))     

if __name__ == '__main__':
    app.run(debug=True)